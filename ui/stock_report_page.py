# ============================================================
# ui/stock_report_page.py — Stock Reports (Item / Metal / Date wise)
# ============================================================
from __future__ import annotations

import csv
import os
from datetime import date as _date

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QPushButton, QTableWidget, QTableWidgetItem, QFrame,
    QHeaderView, QAbstractItemView, QScrollArea, QComboBox,
    QDateEdit, QFileDialog, QMessageBox, QGroupBox,
)
from PyQt5.QtCore import Qt, QDate, pyqtSignal
from PyQt5.QtGui import QFont, QColor

from services.stock_entry_service import get_all_entries

_FILTER_GRP = (
    "QGroupBox{font-size:11px;font-weight:bold;color:#555;"
    "border:1px solid #dfe6e9;border-radius:6px;margin-top:8px;padding-top:4px;}"
    "QGroupBox::title{subcontrol-origin:margin;left:10px;padding:0 4px;}"
)
_BTN_APPLY  = ("QPushButton{background:#2980b9;color:white;border-radius:4px;"
               "padding:2px 14px;font-size:11px;font-weight:bold;border:none;}"
               "QPushButton:hover{background:#2471a3;}")
_BTN_PURPLE = ("QPushButton{background:#8e44ad;color:white;border-radius:4px;"
               "padding:2px 12px;font-size:11px;border:none;}"
               "QPushButton:hover{background:#7d3c98;}")
_BTN_GRAY   = ("QPushButton{background:#7f8c8d;color:white;border-radius:4px;"
               "padding:2px 12px;font-size:11px;border:none;}"
               "QPushButton:hover{background:#626567;}")
_BTN_GREEN  = ("QPushButton{background:#27ae60;color:white;border-radius:4px;"
               "padding:8px 16px;font-size:12px;border:none;}"
               "QPushButton:hover{background:#229954;}")
_BTN_TEAL   = ("QPushButton{background:#16a085;color:white;border-radius:4px;"
               "padding:8px 16px;font-size:12px;border:none;}"
               "QPushButton:hover{background:#138d75;}")
_LINK_BTN   = ("QPushButton{background:#2980b9;color:white;border-radius:3px;"
               "padding:3px 8px;font-size:10px;border:none;font-weight:bold;}"
               "QPushButton:hover{background:#2471a3;}")


def _mk_lbl(text):
    l = QLabel(text)
    l.setStyleSheet("color:#444; font-size:11px; font-weight:600;")
    return l


def _add_date_row(fv, obj, apply_cb, from_start_cb, all_cb):
    """Build the date-range filter row and attach widgets to obj."""
    r1 = QHBoxLayout(); r1.setSpacing(8)
    r1.addWidget(_mk_lbl("From:"))
    obj.dt_from = QDateEdit(); obj.dt_from.setCalendarPopup(True)
    obj.dt_from.setDisplayFormat("dd-MM-yyyy")
    today = QDate.currentDate()
    obj.dt_from.setDate(QDate(today.year(), today.month(), 1))
    obj.dt_from.setFixedHeight(30); r1.addWidget(obj.dt_from)
    r1.addWidget(_mk_lbl("To:"))
    obj.dt_to = QDateEdit(); obj.dt_to.setCalendarPopup(True)
    obj.dt_to.setDisplayFormat("dd-MM-yyyy")
    obj.dt_to.setDate(today); obj.dt_to.setFixedHeight(30); r1.addWidget(obj.dt_to)
    for txt, style, slot in [
        ("Apply Range", _BTN_APPLY,  apply_cb),
        ("From Start",  _BTN_PURPLE, from_start_cb),
        ("All Dates",   _BTN_GRAY,   all_cb),
    ]:
        b = QPushButton(txt); b.setFixedHeight(30); b.setStyleSheet(style)
        b.clicked.connect(slot); r1.addWidget(b)
    r1.addStretch()
    fv.addLayout(r1)


def _make_summary_bar(root, *label_attrs):
    """Add a horizontal summary bar; returns list of QLabel refs in order."""
    sf = QFrame()
    sf.setStyleSheet(
        "background:white; border:1px solid #e0e0e0; border-radius:5px; padding:6px;"
    )
    sl = QHBoxLayout(sf)
    sl.setContentsMargins(10, 4, 10, 4)
    lbls = []
    for i, _ in enumerate(label_attrs):
        lbl = QLabel("—")
        lbl.setFont(QFont("Segoe UI", 11, QFont.Bold))
        lbl.setStyleSheet("background:transparent; color:#2c3e50;")
        if i > 0:
            sl.addStretch()
        sl.addWidget(lbl)
        lbls.append(lbl)
    root.addWidget(sf)
    return lbls


def _footer_row(tbl, vals):
    r = tbl.rowCount(); tbl.insertRow(r)
    fb = QColor("#eaf2ff")
    for c, v in enumerate(vals):
        cell = QTableWidgetItem(str(v))
        cell.setTextAlignment(Qt.AlignCenter)
        cell.setBackground(fb)
        f2 = cell.font(); f2.setBold(True); cell.setFont(f2)
        tbl.setItem(r, c, cell)


def _date_bounds(obj):
    if obj._date_mode == "range":
        return obj.dt_from.date().toString("yyyy-MM-dd"), obj.dt_to.date().toString("yyyy-MM-dd")
    if obj._date_mode == "from_start":
        return "", obj.dt_to.date().toString("yyyy-MM-dd")
    return "", ""


# ============================================================
#  Item-wise
# ============================================================
_IW_LABELS = [
    "Item Name", "Purity", "Metal",
    "Wt In (g)", "Wt Out (g)", "Closing (g)",
    "IN Pcs", "Out Pcs", "Stock Pcs",
]
_IW_RIGHT     = {3, 4, 5, 6, 7, 8}
_IW_SORT_KEYS = {
    0: "item_name", 1: "purity",    2: "metal_type",
    3: "wt_in",     4: "wt_out",    5: "closing_wt",
    6: "in_pcs",    7: "out_pcs",   8: "stock_pcs",
}


class ItemwiseStockPage(QWidget):
    view_datewise_requested = pyqtSignal(str, str, str)  # item_name, purity, metal

    def __init__(self):
        super().__init__()
        self._all_entries = []
        self._rows        = []
        self._sort_col    = -1
        self._sort_asc    = True
        self._date_mode   = "all"
        self._build_ui()

    def _build_ui(self):
        scroll = QScrollArea(self); scroll.setWidgetResizable(True); scroll.setFrameShape(QFrame.NoFrame)
        outer = QVBoxLayout(self); outer.setContentsMargins(0, 0, 0, 0); outer.addWidget(scroll)
        container = QWidget(); scroll.setWidget(container)
        root = QVBoxLayout(container); root.setContentsMargins(25, 20, 25, 20); root.setSpacing(12)

        title = QLabel("📦  Stock Report — Item-wise")
        title.setFont(QFont("Segoe UI", 16, QFont.Bold))
        root.addWidget(title)

        fgrp = QGroupBox("Filters"); fgrp.setStyleSheet(_FILTER_GRP)
        fv = QVBoxLayout(fgrp); fv.setSpacing(6); fv.setContentsMargins(10, 8, 10, 10)
        _add_date_row(fv, self, self._apply_range, self._apply_from_start, self._apply_all)

        r2 = QHBoxLayout(); r2.setSpacing(8)
        r2.addWidget(_mk_lbl("Metal:"))
        self.cmb_metal = QComboBox(); self.cmb_metal.setFixedHeight(30); self.cmb_metal.setMinimumWidth(110)
        self.cmb_metal.currentTextChanged.connect(self._on_metal_changed); r2.addWidget(self.cmb_metal)
        r2.addWidget(_mk_lbl("Purity:"))
        self.cmb_purity = QComboBox(); self.cmb_purity.setFixedHeight(30); self.cmb_purity.setMinimumWidth(86)
        self.cmb_purity.currentTextChanged.connect(self._compute); r2.addWidget(self.cmb_purity)
        r2.addWidget(_mk_lbl("Search:"))
        self.txt_search = QLineEdit(); self.txt_search.setPlaceholderText("Item name…")
        self.txt_search.setFixedHeight(30); self.txt_search.setMinimumWidth(150)
        self.txt_search.textChanged.connect(self._compute); r2.addWidget(self.txt_search)
        r2.addStretch(); fv.addLayout(r2); root.addWidget(fgrp)

        # Table — extra col for Action button
        self.tbl = QTableWidget()
        self.tbl.setColumnCount(len(_IW_LABELS) + 1)
        self.tbl.setHorizontalHeaderLabels([l + " ▲▼" for l in _IW_LABELS] + ["Action"])
        hdr = self.tbl.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.Stretch)
        for c in range(1, len(_IW_LABELS)):
            hdr.setSectionResizeMode(c, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(len(_IW_LABELS), QHeaderView.Fixed)
        self.tbl.setColumnWidth(len(_IW_LABELS), 110)
        hdr.setCursor(Qt.PointingHandCursor); hdr.sectionClicked.connect(self._sort_by)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl.setAlternatingRowColors(True); self.tbl.setMinimumHeight(380)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.verticalHeader().setDefaultSectionSize(32)
        root.addWidget(self.tbl)

        self._sum_lbls = _make_summary_bar(root, "items", "wt_in", "wt_out", "closing")

        act = QHBoxLayout(); act.addStretch()
        for txt, style, slot in [
            ("📤  Export CSV",   _BTN_GREEN, self._export_csv),
            ("📊  Export Excel", _BTN_TEAL,  self._export_excel),
        ]:
            b = QPushButton(txt); b.setStyleSheet(style); b.clicked.connect(slot); act.addWidget(b)
        root.addLayout(act)

    def _apply_range(self):      self._date_mode = "range";       self._compute()
    def _apply_from_start(self): self._date_mode = "from_start";  self._compute()
    def _apply_all(self):        self._date_mode = "all";          self._compute()

    def _rebuild_combos(self):
        metals = sorted({e.get("metal_type","") for e in self._all_entries if e.get("metal_type")})
        curr_m = self.cmb_metal.currentText()
        self.cmb_metal.blockSignals(True); self.cmb_metal.clear()
        self.cmb_metal.addItem("All"); self.cmb_metal.addItems(metals)
        self.cmb_metal.setCurrentIndex(max(self.cmb_metal.findText(curr_m), 0))
        self.cmb_metal.blockSignals(False)
        self._update_purity(self.cmb_metal.currentText())

    def _update_purity(self, metal):
        if metal and metal != "All":
            purities = sorted({e.get("purity","") for e in self._all_entries
                               if e.get("metal_type","") == metal and e.get("purity")})
        else:
            purities = sorted({e.get("purity","") for e in self._all_entries if e.get("purity")})
        curr = self.cmb_purity.currentText()
        self.cmb_purity.blockSignals(True); self.cmb_purity.clear()
        self.cmb_purity.addItem("All"); self.cmb_purity.addItems(purities)
        self.cmb_purity.setCurrentIndex(max(self.cmb_purity.findText(curr), 0))
        self.cmb_purity.blockSignals(False)

    def _on_metal_changed(self, metal):
        self._update_purity(metal); self._compute()

    def refresh(self):
        self._all_entries = get_all_entries()
        self._rebuild_combos(); self._compute()

    def _compute(self):
        from_s, to_s = _date_bounds(self)
        metal  = self.cmb_metal.currentText()
        purity = self.cmb_purity.currentText()
        item_q = self.txt_search.text().strip().lower()

        before, period = [], []
        for e in self._all_entries:
            if metal  != "All" and e.get("metal_type","") != metal:   continue
            if purity != "All" and e.get("purity","")     != purity:  continue
            if item_q and item_q not in e.get("item_name","").lower(): continue
            d = e.get("voucher_date","")
            if from_s and d < from_s:  before.append(e)
            elif to_s and d > to_s:    pass
            else:                      period.append(e)

        groups = {}
        def _get(e):
            key = (e.get("item_name",""), e.get("purity",""), e.get("metal_type",""))
            if key not in groups:
                groups[key] = {"item_name": key[0], "purity": key[1], "metal_type": key[2],
                               "open_wt": 0.0, "open_pcs": 0,
                               "wt_in": 0.0, "in_pcs": 0, "wt_out": 0.0, "out_pcs": 0}
            return groups[key]

        for e in before:
            g = _get(e)
            if e.get("entry_type") == "IN":
                g["open_wt"]  += float(e.get("net_wt",    0) or 0)
                g["open_pcs"] += int(e.get("qty_in",      0) or 0)
            else:
                g["open_wt"]  -= float(e.get("out_net_wt", 0) or 0)
                g["open_pcs"] -= int(e.get("qty_out",      0) or 0)
        for e in period:
            g = _get(e)
            if e.get("entry_type") == "IN":
                g["wt_in"]  += float(e.get("net_wt",    0) or 0)
                g["in_pcs"] += int(e.get("qty_in",      0) or 0)
            else:
                g["wt_out"]  += float(e.get("out_net_wt", 0) or 0)
                g["out_pcs"] += int(e.get("qty_out",      0) or 0)

        rows = []
        for g in groups.values():
            g["closing_wt"] = round(g["open_wt"] + g["wt_in"] - g["wt_out"], 3)
            g["stock_pcs"]  = g["open_pcs"] + g["in_pcs"] - g["out_pcs"]
            rows.append(g)

        self._rows = self._apply_sort(rows)
        self._populate(self._rows)

    def _sort_by(self, col):
        if col == len(_IW_LABELS): return
        if self._sort_col == col:
            if not self._sort_asc:
                self._sort_asc = True
            else:
                self._sort_col = -1; self._refresh_headers(); self._populate(self._rows); return
        else:
            self._sort_col = col; self._sort_asc = False
        self._refresh_headers(); self._rows = self._apply_sort(self._rows); self._populate(self._rows)

    def _apply_sort(self, data):
        if self._sort_col < 0 or self._sort_col not in _IW_SORT_KEYS:
            return sorted(data, key=lambda x: x.get("item_name","").lower())
        k = _IW_SORT_KEYS[self._sort_col]
        def _key(r):
            v = r.get(k, "")
            try:   return (0, float(v))
            except: return (1, str(v).lower())
        return sorted(data, key=_key, reverse=not self._sort_asc)

    def _refresh_headers(self):
        all_hdrs = _IW_LABELS + ["Action"]
        for i, lbl in enumerate(all_hdrs):
            item = self.tbl.horizontalHeaderItem(i)
            if not item: continue
            if i == len(_IW_LABELS):    item.setText("Action")
            elif i == self._sort_col:   item.setText(lbl + (" ▲" if self._sort_asc else " ▼"))
            else:                       item.setText(lbl + " ▲▼")

    def _populate(self, data):
        self.tbl.setRowCount(0)
        t_in = t_out = t_close = 0.0
        for g in data:
            r = self.tbl.rowCount(); self.tbl.insertRow(r)
            closing = g["closing_wt"]; stock = g["stock_pcs"]
            vals = [g.get("item_name",""), g.get("purity",""), g.get("metal_type",""),
                    f"{g['wt_in']:.3f}", f"{g['wt_out']:.3f}", f"{closing:.3f}",
                    str(g["in_pcs"]), str(g["out_pcs"]), str(stock)]
            for c, v in enumerate(vals):
                cell = QTableWidgetItem(v)
                cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter if c in _IW_RIGHT
                                      else Qt.AlignLeft | Qt.AlignVCenter)
                if c == 5: cell.setForeground(QColor("#1e8449") if closing > 0 else QColor("#e74c3c"))
                if c == 8 and stock <= 0: cell.setForeground(QColor("#e74c3c"))
                self.tbl.setItem(r, c, cell)
            btn = QPushButton("📅 Date-wise"); btn.setStyleSheet(_LINK_BTN)
            nm, pu, mt = g.get("item_name",""), g.get("purity",""), g.get("metal_type","")
            btn.clicked.connect(lambda _, n=nm, p=pu, m=mt: self.view_datewise_requested.emit(n, p, m))
            self.tbl.setCellWidget(r, len(_IW_LABELS), btn)
            t_in += g["wt_in"]; t_out += g["wt_out"]; t_close += closing

        if data:
            _footer_row(self.tbl, ["TOTAL","","", f"{t_in:.3f}", f"{t_out:.3f}", f"{t_close:.3f}","","","",""])

        self._sum_lbls[0].setText(f"Items: {len(data)}")
        self._sum_lbls[1].setText(f"Wt IN: {t_in:.3f} g")
        self._sum_lbls[2].setText(f"Wt OUT: {t_out:.3f} g")
        self._sum_lbls[3].setText(f"Closing: {t_close:.3f} g")

    def _export_csv(self):
        if not self._rows: QMessageBox.information(self, "Export", "No data."); return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export CSV", f"stock_itemwise_{_date.today()}.csv", "CSV (*.csv)")
        if not path: return
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["Item Name","Purity","Metal","Wt In (g)","Wt Out (g)","Closing (g)","IN Pcs","Out Pcs","Stock Pcs"])
            for g in self._rows:
                w.writerow([g.get("item_name",""), g.get("purity",""), g.get("metal_type",""),
                            f"{g.get('wt_in',0):.3f}", f"{g.get('wt_out',0):.3f}", f"{g.get('closing_wt',0):.3f}",
                            g.get("in_pcs",0), g.get("out_pcs",0), g.get("stock_pcs",0)])
        QMessageBox.information(self, "Saved", f"Saved to:\n{path}")
        try: os.startfile(path)
        except: pass

    def _export_excel(self):
        if not _HAS_OPENPYXL: QMessageBox.warning(self, "Export", "openpyxl not installed."); return
        if not self._rows:    QMessageBox.information(self, "Export", "No data."); return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Excel", f"stock_itemwise_{_date.today()}.xlsx", "Excel (*.xlsx)")
        if not path: return
        try:
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Item-wise Stock"
            hdrs = ["Item Name","Purity","Metal","Wt In (g)","Wt Out (g)","Closing (g)","IN Pcs","Out Pcs","Stock Pcs"]
            ws.append(hdrs)
            hf = PatternFill("solid", fgColor="2C3E50"); hfont = Font(bold=True, color="FFFFFF")
            for cell in ws[1]: cell.fill = hf; cell.font = hfont; cell.alignment = Alignment(horizontal="center")
            _money = {4,5,6}
            for g in self._rows:
                ws.append([g.get("item_name",""), g.get("purity",""), g.get("metal_type",""),
                           round(g.get("wt_in",0),3), round(g.get("wt_out",0),3), round(g.get("closing_wt",0),3),
                           g.get("in_pcs",0), g.get("out_pcs",0), g.get("stock_pcs",0)])
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    if cell.column in _money: cell.number_format = "#,##0.000"
            ws.column_dimensions["A"].width = 28
            for col_l in ["B","C","D","E","F","G","H","I"]: ws.column_dimensions[col_l].width = 13
            wb.save(path)
            QMessageBox.information(self, "Saved", f"Excel saved to:\n{path}")
            try: os.startfile(path)
            except: pass
        except Exception as exc:
            QMessageBox.critical(self, "Export Error", str(exc))


# ============================================================
#  Metal-wise
# ============================================================
_MW_LABELS    = ["Metal", "Wt In (g)", "Wt Out (g)", "Closing (g)", "IN Pcs", "Out Pcs", "Stock Pcs"]
_MW_RIGHT     = {1, 2, 3, 4, 5, 6}
_MW_SORT_KEYS = {0:"metal_type", 1:"wt_in", 2:"wt_out", 3:"closing_wt", 4:"in_pcs", 5:"out_pcs", 6:"stock_pcs"}


class MetalwiseStockPage(QWidget):
    def __init__(self):
        super().__init__()
        self._all_entries = []
        self._rows        = []
        self._sort_col    = -1
        self._sort_asc    = True
        self._date_mode   = "all"
        self._build_ui()

    def _build_ui(self):
        scroll = QScrollArea(self); scroll.setWidgetResizable(True); scroll.setFrameShape(QFrame.NoFrame)
        outer = QVBoxLayout(self); outer.setContentsMargins(0, 0, 0, 0); outer.addWidget(scroll)
        container = QWidget(); scroll.setWidget(container)
        root = QVBoxLayout(container); root.setContentsMargins(25, 20, 25, 20); root.setSpacing(12)

        title = QLabel("🔩  Stock Report — Metal-wise")
        title.setFont(QFont("Segoe UI", 16, QFont.Bold))
        root.addWidget(title)

        fgrp = QGroupBox("Filters"); fgrp.setStyleSheet(_FILTER_GRP)
        fv = QVBoxLayout(fgrp); fv.setSpacing(6); fv.setContentsMargins(10, 8, 10, 10)
        _add_date_row(fv, self, self._apply_range, self._apply_from_start, self._apply_all)
        root.addWidget(fgrp)

        self.tbl = QTableWidget()
        self.tbl.setColumnCount(len(_MW_LABELS))
        self.tbl.setHorizontalHeaderLabels([l + " ▲▼" for l in _MW_LABELS])
        hdr = self.tbl.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.Stretch)
        for c in range(1, len(_MW_LABELS)): hdr.setSectionResizeMode(c, QHeaderView.ResizeToContents)
        hdr.setCursor(Qt.PointingHandCursor); hdr.sectionClicked.connect(self._sort_by)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl.setAlternatingRowColors(True); self.tbl.setMinimumHeight(300)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.verticalHeader().setDefaultSectionSize(32)
        root.addWidget(self.tbl)

        self._sum_lbls = _make_summary_bar(root, "metals", "wt_in", "wt_out", "closing")

        act = QHBoxLayout(); act.addStretch()
        for txt, style, slot in [
            ("📤  Export CSV",   _BTN_GREEN, self._export_csv),
            ("📊  Export Excel", _BTN_TEAL,  self._export_excel),
        ]:
            b = QPushButton(txt); b.setStyleSheet(style); b.clicked.connect(slot); act.addWidget(b)
        root.addLayout(act)

    def _apply_range(self):      self._date_mode = "range";       self._compute()
    def _apply_from_start(self): self._date_mode = "from_start";  self._compute()
    def _apply_all(self):        self._date_mode = "all";          self._compute()

    def refresh(self):
        self._all_entries = get_all_entries(); self._compute()

    def _compute(self):
        from_s, to_s = _date_bounds(self)
        before, period = [], []
        for e in self._all_entries:
            if not e.get("metal_type"): continue
            d = e.get("voucher_date","")
            if from_s and d < from_s:  before.append(e)
            elif to_s and d > to_s:    pass
            else:                      period.append(e)

        groups = {}
        def _get(e):
            key = e.get("metal_type","") or "Unknown"
            if key not in groups:
                groups[key] = {"metal_type": key, "open_wt": 0.0, "open_pcs": 0,
                               "wt_in": 0.0, "in_pcs": 0, "wt_out": 0.0, "out_pcs": 0}
            return groups[key]

        for e in before:
            g = _get(e)
            if e.get("entry_type") == "IN":
                g["open_wt"]  += float(e.get("net_wt",    0) or 0)
                g["open_pcs"] += int(e.get("qty_in",      0) or 0)
            else:
                g["open_wt"]  -= float(e.get("out_net_wt", 0) or 0)
                g["open_pcs"] -= int(e.get("qty_out",      0) or 0)
        for e in period:
            g = _get(e)
            if e.get("entry_type") == "IN":
                g["wt_in"]  += float(e.get("net_wt",    0) or 0)
                g["in_pcs"] += int(e.get("qty_in",      0) or 0)
            else:
                g["wt_out"]  += float(e.get("out_net_wt", 0) or 0)
                g["out_pcs"] += int(e.get("qty_out",      0) or 0)

        rows = []
        for g in groups.values():
            g["closing_wt"] = round(g["open_wt"] + g["wt_in"] - g["wt_out"], 3)
            g["stock_pcs"]  = g["open_pcs"] + g["in_pcs"] - g["out_pcs"]
            rows.append(g)

        self._rows = self._apply_sort(rows)
        self._populate(self._rows)

    def _sort_by(self, col):
        if self._sort_col == col:
            if not self._sort_asc:
                self._sort_asc = True
            else:
                self._sort_col = -1; self._refresh_headers(); self._populate(self._rows); return
        else:
            self._sort_col = col; self._sort_asc = False
        self._refresh_headers(); self._rows = self._apply_sort(self._rows); self._populate(self._rows)

    def _apply_sort(self, data):
        if self._sort_col < 0 or self._sort_col not in _MW_SORT_KEYS:
            return sorted(data, key=lambda x: x.get("metal_type","").lower())
        k = _MW_SORT_KEYS[self._sort_col]
        def _key(r):
            v = r.get(k, "")
            try:    return (0, float(v))
            except: return (1, str(v).lower())
        return sorted(data, key=_key, reverse=not self._sort_asc)

    def _refresh_headers(self):
        for i, lbl in enumerate(_MW_LABELS):
            item = self.tbl.horizontalHeaderItem(i)
            if not item: continue
            item.setText(lbl + (" ▲" if self._sort_asc else " ▼") if i == self._sort_col else lbl + " ▲▼")

    def _populate(self, data):
        self.tbl.setRowCount(0)
        t_in = t_out = t_close = 0.0
        for g in data:
            r = self.tbl.rowCount(); self.tbl.insertRow(r)
            closing = g["closing_wt"]; stock = g["stock_pcs"]
            vals = [g.get("metal_type",""), f"{g['wt_in']:.3f}", f"{g['wt_out']:.3f}", f"{closing:.3f}",
                    str(g["in_pcs"]), str(g["out_pcs"]), str(stock)]
            for c, v in enumerate(vals):
                cell = QTableWidgetItem(v)
                cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter if c in _MW_RIGHT
                                      else Qt.AlignLeft | Qt.AlignVCenter)
                if c == 3: cell.setForeground(QColor("#1e8449") if closing > 0 else QColor("#e74c3c"))
                if c == 6 and stock <= 0: cell.setForeground(QColor("#e74c3c"))
                self.tbl.setItem(r, c, cell)
            t_in += g["wt_in"]; t_out += g["wt_out"]; t_close += closing

        if data:
            _footer_row(self.tbl, ["TOTAL", f"{t_in:.3f}", f"{t_out:.3f}", f"{t_close:.3f}", "", "", ""])

        self._sum_lbls[0].setText(f"Metals: {len(data)}")
        self._sum_lbls[1].setText(f"Wt IN: {t_in:.3f} g")
        self._sum_lbls[2].setText(f"Wt OUT: {t_out:.3f} g")
        self._sum_lbls[3].setText(f"Closing: {t_close:.3f} g")

    def _export_csv(self):
        if not self._rows: QMessageBox.information(self, "Export", "No data."); return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export CSV", f"stock_metalwise_{_date.today()}.csv", "CSV (*.csv)")
        if not path: return
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["Metal","Wt In (g)","Wt Out (g)","Closing (g)","IN Pcs","Out Pcs","Stock Pcs"])
            for g in self._rows:
                w.writerow([g.get("metal_type",""),
                            f"{g.get('wt_in',0):.3f}", f"{g.get('wt_out',0):.3f}", f"{g.get('closing_wt',0):.3f}",
                            g.get("in_pcs",0), g.get("out_pcs",0), g.get("stock_pcs",0)])
        QMessageBox.information(self, "Saved", f"Saved to:\n{path}")
        try: os.startfile(path)
        except: pass

    def _export_excel(self):
        if not _HAS_OPENPYXL: QMessageBox.warning(self, "Export", "openpyxl not installed."); return
        if not self._rows:    QMessageBox.information(self, "Export", "No data."); return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Excel", f"stock_metalwise_{_date.today()}.xlsx", "Excel (*.xlsx)")
        if not path: return
        try:
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Metal-wise Stock"
            hdrs = ["Metal","Wt In (g)","Wt Out (g)","Closing (g)","IN Pcs","Out Pcs","Stock Pcs"]
            ws.append(hdrs)
            hf = PatternFill("solid", fgColor="2C3E50"); hfont = Font(bold=True, color="FFFFFF")
            for cell in ws[1]: cell.fill = hf; cell.font = hfont; cell.alignment = Alignment(horizontal="center")
            _money = {2,3,4}
            for g in self._rows:
                ws.append([g.get("metal_type",""),
                           round(g.get("wt_in",0),3), round(g.get("wt_out",0),3), round(g.get("closing_wt",0),3),
                           g.get("in_pcs",0), g.get("out_pcs",0), g.get("stock_pcs",0)])
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    if cell.column in _money: cell.number_format = "#,##0.000"
            ws.column_dimensions["A"].width = 18
            for col_l in ["B","C","D","E","F","G"]: ws.column_dimensions[col_l].width = 14
            wb.save(path)
            QMessageBox.information(self, "Saved", f"Excel saved to:\n{path}")
            try: os.startfile(path)
            except: pass
        except Exception as exc:
            QMessageBox.critical(self, "Export Error", str(exc))


# ============================================================
#  Date-wise
# ============================================================
_DW_LABELS    = ["Date", "Wt In (g)", "Wt Out (g)", "Closing (g)", "IN Pcs", "Out Pcs", "Stock Pcs"]
_DW_RIGHT     = {1, 2, 3, 4, 5, 6}
_DW_SORT_KEYS = {0:"date", 1:"wt_in", 2:"wt_out", 3:"closing_wt", 4:"in_pcs", 5:"out_pcs", 6:"stock_pcs"}


class DatewiseStockPage(QWidget):
    def __init__(self):
        super().__init__()
        self._all_entries    = []
        self._rows           = []
        self._sort_col       = -1
        self._sort_asc       = True
        self._date_mode      = "all"
        self._filter_item    = ""
        self._filter_purity  = ""
        self._filter_metal   = ""
        self._final_closing  = 0.0
        self._build_ui()

    def _build_ui(self):
        scroll = QScrollArea(self); scroll.setWidgetResizable(True); scroll.setFrameShape(QFrame.NoFrame)
        outer = QVBoxLayout(self); outer.setContentsMargins(0, 0, 0, 0); outer.addWidget(scroll)
        container = QWidget(); scroll.setWidget(container)
        root = QVBoxLayout(container); root.setContentsMargins(25, 20, 25, 20); root.setSpacing(12)

        title = QLabel("📅  Stock Report — Date-wise")
        title.setFont(QFont("Segoe UI", 16, QFont.Bold))
        root.addWidget(title)

        # Item filter banner
        self._banner = QFrame()
        self._banner.setStyleSheet("QFrame{background:#2c3e50; border-radius:5px; padding:2px;}")
        bl = QHBoxLayout(self._banner); bl.setContentsMargins(12, 6, 12, 6)
        self._banner_lbl = QLabel("Showing: All Items")
        self._banner_lbl.setStyleSheet("color:white; font-size:12px; font-weight:bold; background:transparent;")
        self._btn_clear = QPushButton("✕  Show All Items")
        self._btn_clear.setStyleSheet(
            "QPushButton{background:#e74c3c;color:white;border-radius:3px;"
            "padding:3px 10px;font-size:11px;border:none;}"
            "QPushButton:hover{background:#c0392b;}"
        )
        self._btn_clear.clicked.connect(self._clear_item_filter)
        self._btn_clear.setVisible(False)
        bl.addWidget(self._banner_lbl); bl.addStretch(); bl.addWidget(self._btn_clear)
        root.addWidget(self._banner)

        fgrp = QGroupBox("Filters"); fgrp.setStyleSheet(_FILTER_GRP)
        fv = QVBoxLayout(fgrp); fv.setSpacing(6); fv.setContentsMargins(10, 8, 10, 10)
        _add_date_row(fv, self, self._apply_range, self._apply_from_start, self._apply_all)

        r2 = QHBoxLayout(); r2.setSpacing(8)
        r2.addWidget(_mk_lbl("Metal:"))
        self.cmb_metal = QComboBox(); self.cmb_metal.setFixedHeight(30); self.cmb_metal.setMinimumWidth(110)
        self.cmb_metal.currentTextChanged.connect(self._on_metal_changed); r2.addWidget(self.cmb_metal)
        r2.addWidget(_mk_lbl("Purity:"))
        self.cmb_purity = QComboBox(); self.cmb_purity.setFixedHeight(30); self.cmb_purity.setMinimumWidth(86)
        self.cmb_purity.currentTextChanged.connect(self._compute); r2.addWidget(self.cmb_purity)
        r2.addWidget(_mk_lbl("Item:"))
        self.cmb_item = QComboBox(); self.cmb_item.setFixedHeight(30); self.cmb_item.setMinimumWidth(180)
        self.cmb_item.currentTextChanged.connect(self._on_item_changed); r2.addWidget(self.cmb_item)
        r2.addStretch(); fv.addLayout(r2); root.addWidget(fgrp)

        self.tbl = QTableWidget()
        self.tbl.setColumnCount(len(_DW_LABELS))
        self.tbl.setHorizontalHeaderLabels([l + " ▲▼" for l in _DW_LABELS])
        hdr = self.tbl.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.Stretch)
        for c in range(1, len(_DW_LABELS)): hdr.setSectionResizeMode(c, QHeaderView.ResizeToContents)
        hdr.setCursor(Qt.PointingHandCursor); hdr.sectionClicked.connect(self._sort_by)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl.setAlternatingRowColors(True); self.tbl.setMinimumHeight(380)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.verticalHeader().setDefaultSectionSize(32)
        root.addWidget(self.tbl)

        self._sum_lbls = _make_summary_bar(root, "days", "wt_in", "wt_out", "closing")

        act = QHBoxLayout(); act.addStretch()
        for txt, style, slot in [
            ("📤  Export CSV",   _BTN_GREEN, self._export_csv),
            ("📊  Export Excel", _BTN_TEAL,  self._export_excel),
        ]:
            b = QPushButton(txt); b.setStyleSheet(style); b.clicked.connect(slot); act.addWidget(b)
        root.addLayout(act)

    # ── Item filter (set from ItemwisePage via dashboard) ────
    def set_item_filter(self, item_name, purity, metal):
        self._filter_item   = item_name
        self._filter_purity = purity
        self._filter_metal  = metal
        self._update_banner()
        # Sync combos silently
        for cmb, val in [(self.cmb_item, item_name), (self.cmb_metal, metal), (self.cmb_purity, purity)]:
            idx = cmb.findText(val)
            if idx >= 0:
                cmb.blockSignals(True); cmb.setCurrentIndex(idx); cmb.blockSignals(False)
        self._compute()

    def _clear_item_filter(self):
        self._filter_item = self._filter_purity = self._filter_metal = ""
        self._update_banner()
        self.cmb_item.blockSignals(True); self.cmb_item.setCurrentIndex(0); self.cmb_item.blockSignals(False)
        self._compute()

    def _update_banner(self):
        if self._filter_item:
            parts = [self._filter_item]
            if self._filter_purity: parts.append(f"Purity: {self._filter_purity}")
            if self._filter_metal:  parts.append(f"Metal: {self._filter_metal}")
            self._banner_lbl.setText("Showing: " + "  |  ".join(parts))
            self._banner.setStyleSheet("QFrame{background:#1a3a5c; border-radius:5px; padding:2px;}")
            self._btn_clear.setVisible(True)
        else:
            self._banner_lbl.setText("Showing: All Items")
            self._banner.setStyleSheet("QFrame{background:#2c3e50; border-radius:5px; padding:2px;}")
            self._btn_clear.setVisible(False)

    def _apply_range(self):      self._date_mode = "range";       self._compute()
    def _apply_from_start(self): self._date_mode = "from_start";  self._compute()
    def _apply_all(self):        self._date_mode = "all";          self._compute()

    def _rebuild_combos(self):
        metals = sorted({e.get("metal_type","") for e in self._all_entries if e.get("metal_type")})
        items  = sorted({e.get("item_name","")  for e in self._all_entries if e.get("item_name")})

        for cmb, vals, curr in [
            (self.cmb_metal, metals, self.cmb_metal.currentText()),
            (self.cmb_item,  items,  self.cmb_item.currentText()),
        ]:
            cmb.blockSignals(True); cmb.clear(); cmb.addItem("All"); cmb.addItems(vals)
            cmb.setCurrentIndex(max(cmb.findText(curr), 0)); cmb.blockSignals(False)
        self._update_purity(self.cmb_metal.currentText())

    def _update_purity(self, metal):
        if metal and metal != "All":
            purities = sorted({e.get("purity","") for e in self._all_entries
                               if e.get("metal_type","") == metal and e.get("purity")})
        else:
            purities = sorted({e.get("purity","") for e in self._all_entries if e.get("purity")})
        curr = self.cmb_purity.currentText()
        self.cmb_purity.blockSignals(True); self.cmb_purity.clear()
        self.cmb_purity.addItem("All"); self.cmb_purity.addItems(purities)
        self.cmb_purity.setCurrentIndex(max(self.cmb_purity.findText(curr), 0))
        self.cmb_purity.blockSignals(False)

    def _on_metal_changed(self, metal):
        self._update_purity(metal); self._compute()

    def _on_item_changed(self, item_name):
        if item_name == "All":
            self._filter_item = self._filter_purity = self._filter_metal = ""
        else:
            self._filter_item = item_name
        self._update_banner(); self._compute()

    def refresh(self):
        self._all_entries = get_all_entries()
        self._rebuild_combos(); self._compute()

    def _compute(self):
        from_s, to_s = _date_bounds(self)
        metal  = self.cmb_metal.currentText()
        purity = self.cmb_purity.currentText()
        item_f = self._filter_item or (
            self.cmb_item.currentText() if self.cmb_item.currentText() != "All" else "")

        before, period = [], []
        for e in self._all_entries:
            if metal  != "All" and e.get("metal_type","") != metal:  continue
            if purity != "All" and e.get("purity","")     != purity: continue
            if item_f and e.get("item_name","") != item_f:           continue
            d = e.get("voucher_date","")
            if from_s and d < from_s:  before.append(e)
            elif to_s and d > to_s:    pass
            else:                      period.append(e)

        # Opening balance before range
        open_wt = 0.0; open_pcs = 0
        for e in before:
            if e.get("entry_type") == "IN":
                open_wt  += float(e.get("net_wt",    0) or 0)
                open_pcs += int(e.get("qty_in",      0) or 0)
            else:
                open_wt  -= float(e.get("out_net_wt", 0) or 0)
                open_pcs -= int(e.get("qty_out",      0) or 0)

        # Aggregate by date
        by_date = {}
        for e in period:
            d = e.get("voucher_date","") or "Unknown"
            if d not in by_date:
                by_date[d] = {"date": d, "wt_in": 0.0, "wt_out": 0.0, "in_pcs": 0, "out_pcs": 0}
            g = by_date[d]
            if e.get("entry_type") == "IN":
                g["wt_in"]  += float(e.get("net_wt",    0) or 0)
                g["in_pcs"] += int(e.get("qty_in",      0) or 0)
            else:
                g["wt_out"]  += float(e.get("out_net_wt", 0) or 0)
                g["out_pcs"] += int(e.get("qty_out",      0) or 0)

        # Build rows with cumulative closing (always sorted by date for correctness)
        rows = []
        run_wt = open_wt; run_pcs = open_pcs
        for d in sorted(by_date.keys()):
            g = by_date[d]
            run_wt  = round(run_wt  + g["wt_in"]  - g["wt_out"],  3)
            run_pcs = run_pcs + g["in_pcs"] - g["out_pcs"]
            rows.append({"date": d, "wt_in": g["wt_in"], "wt_out": g["wt_out"],
                         "closing_wt": run_wt, "in_pcs": g["in_pcs"],
                         "out_pcs": g["out_pcs"], "stock_pcs": run_pcs})

        self._final_closing = rows[-1]["closing_wt"] if rows else round(open_wt, 3)
        self._period_wt_in  = sum(r["wt_in"]  for r in rows)
        self._period_wt_out = sum(r["wt_out"] for r in rows)

        self._rows = self._apply_sort(rows)
        self._populate(self._rows)

    def _sort_by(self, col):
        if self._sort_col == col:
            if not self._sort_asc:
                self._sort_asc = True
            else:
                self._sort_col = -1; self._refresh_headers(); self._populate(self._rows); return
        else:
            self._sort_col = col; self._sort_asc = False
        self._refresh_headers(); self._rows = self._apply_sort(self._rows); self._populate(self._rows)

    def _apply_sort(self, data):
        if self._sort_col < 0 or self._sort_col not in _DW_SORT_KEYS:
            return sorted(data, key=lambda x: x.get("date",""))
        k = _DW_SORT_KEYS[self._sort_col]
        def _key(r):
            v = r.get(k, "")
            try:    return (0, float(v))
            except: return (1, str(v).lower())
        return sorted(data, key=_key, reverse=not self._sort_asc)

    def _refresh_headers(self):
        for i, lbl in enumerate(_DW_LABELS):
            item = self.tbl.horizontalHeaderItem(i)
            if not item: continue
            item.setText(lbl + (" ▲" if self._sort_asc else " ▼") if i == self._sort_col else lbl + " ▲▼")

    def _populate(self, data):
        self.tbl.setRowCount(0)
        t_in = t_out = 0.0; t_in_pcs = t_out_pcs = 0
        for g in data:
            r = self.tbl.rowCount(); self.tbl.insertRow(r)
            closing = g["closing_wt"]; stock = g["stock_pcs"]
            vals = [g.get("date",""), f"{g['wt_in']:.3f}", f"{g['wt_out']:.3f}", f"{closing:.3f}",
                    str(g["in_pcs"]), str(g["out_pcs"]), str(stock)]
            for c, v in enumerate(vals):
                cell = QTableWidgetItem(v)
                cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter if c in _DW_RIGHT
                                      else Qt.AlignLeft | Qt.AlignVCenter)
                if c == 3: cell.setForeground(QColor("#1e8449") if closing > 0 else QColor("#e74c3c"))
                if c == 6 and stock <= 0: cell.setForeground(QColor("#e74c3c"))
                self.tbl.setItem(r, c, cell)
            t_in += g["wt_in"]; t_out += g["wt_out"]
            t_in_pcs += g["in_pcs"]; t_out_pcs += g["out_pcs"]

        if data:
            net = round(t_in - t_out, 3)
            _footer_row(self.tbl, ["TOTAL", f"{t_in:.3f}", f"{t_out:.3f}", f"{net:.3f}",
                                   str(t_in_pcs), str(t_out_pcs), ""])

        self._sum_lbls[0].setText(f"Days: {len(data)}")
        self._sum_lbls[1].setText(f"Wt IN: {self._period_wt_in:.3f} g")
        self._sum_lbls[2].setText(f"Wt OUT: {self._period_wt_out:.3f} g")
        self._sum_lbls[3].setText(f"Closing: {self._final_closing:.3f} g")

    def _item_tag(self):
        return f"_{self._filter_item.replace(' ','_')}" if self._filter_item else ""

    def _export_csv(self):
        if not self._rows: QMessageBox.information(self, "Export", "No data."); return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export CSV",
            f"stock_datewise{self._item_tag()}_{_date.today()}.csv", "CSV (*.csv)")
        if not path: return
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["Date","Wt In (g)","Wt Out (g)","Closing (g)","IN Pcs","Out Pcs","Stock Pcs"])
            for g in self._rows:
                w.writerow([g.get("date",""),
                            f"{g.get('wt_in',0):.3f}", f"{g.get('wt_out',0):.3f}", f"{g.get('closing_wt',0):.3f}",
                            g.get("in_pcs",0), g.get("out_pcs",0), g.get("stock_pcs",0)])
        QMessageBox.information(self, "Saved", f"Saved to:\n{path}")
        try: os.startfile(path)
        except: pass

    def _export_excel(self):
        if not _HAS_OPENPYXL: QMessageBox.warning(self, "Export", "openpyxl not installed."); return
        if not self._rows:    QMessageBox.information(self, "Export", "No data."); return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Excel",
            f"stock_datewise{self._item_tag()}_{_date.today()}.xlsx", "Excel (*.xlsx)")
        if not path: return
        try:
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Date-wise Stock"
            hdrs = ["Date","Wt In (g)","Wt Out (g)","Closing (g)","IN Pcs","Out Pcs","Stock Pcs"]
            ws.append(hdrs)
            hf = PatternFill("solid", fgColor="2C3E50"); hfont = Font(bold=True, color="FFFFFF")
            for cell in ws[1]: cell.fill = hf; cell.font = hfont; cell.alignment = Alignment(horizontal="center")
            _money = {2,3,4}
            for g in self._rows:
                ws.append([g.get("date",""),
                           round(g.get("wt_in",0),3), round(g.get("wt_out",0),3), round(g.get("closing_wt",0),3),
                           g.get("in_pcs",0), g.get("out_pcs",0), g.get("stock_pcs",0)])
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    if cell.column in _money: cell.number_format = "#,##0.000"
            ws.column_dimensions["A"].width = 14
            for col_l in ["B","C","D","E","F","G"]: ws.column_dimensions[col_l].width = 14
            wb.save(path)
            QMessageBox.information(self, "Saved", f"Excel saved to:\n{path}")
            try: os.startfile(path)
            except: pass
        except Exception as exc:
            QMessageBox.critical(self, "Export Error", str(exc))


# Backward-compat alias
StockReportPage = ItemwiseStockPage
