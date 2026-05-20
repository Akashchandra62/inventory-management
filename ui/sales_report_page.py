# ============================================================
# ui/sales_report_page.py - Sales Report / Invoice History
# ============================================================

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QTableWidget, QTableWidgetItem,
    QLineEdit, QDateEdit, QGroupBox, QFrame,
    QHeaderView, QAbstractItemView, QMessageBox,
    QScrollArea, QComboBox,
)
from PyQt5.QtCore import Qt, QDate, pyqtSignal
from PyQt5.QtGui import QFont, QColor
from services.invoice_service import get_all_invoices
from services.metal_service import get_metals
from services.item_catalog_service import get_catalog
from app.utils import format_currency
from app.printer_helper import save_invoice_as_pdf
from app.config import AppConfig
import csv, os
from datetime import date, timedelta


_RPT_LABELS = [
    "Invoice No", "Date", "Customer",
    "Gross Sale (₹)", "Discount (₹)", "Sale Amt (₹)",
    "CGST (₹)", "SGST (₹)", "Other Purch (₹)",
    "Advance (₹)", "Cash (₹)", "Card (₹)",
    "Cheque (₹)", "Dues (₹)",
]
_NON_SORTABLE: set = set()
_SORT_KEYS = {
    0: "invoice_number", 1: "date",        2: "customer_name",
    3: "gross_sale",     4: "discount",     5: "sale_amount",
    6: "cgst",           7: "sgst",         8: "other_purchase",
    9: "advance",        10: "cash",        11: "card",
    12: "cheque",        13: "dues",
}


def _due_amount(inv: dict) -> float:
    return float(inv.get("due_amount", 0) or 0)


def _card_style(bg: str) -> str:
    return f"QFrame {{ background: {bg}; border-radius: 8px; border: none; }}"


def _build_inv_row(inv: dict, catalog_metal_map: dict) -> dict:
    """Build one summary row per invoice."""
    items = inv.get("items", [])
    metals     = {it.get("metal", "").strip().lower() for it in items}
    purities   = {it.get("purity", "").strip().lower() for it in items}
    categories = {it.get("category", "").strip() for it in items}
    # fill metal from catalog if blank
    metals |= {
        catalog_metal_map.get(it.get("name", "").strip().lower(), "").lower()
        for it in items
    }
    metals.discard("")

    return {
        "_inv":            inv,
        "invoice_number":  inv.get("invoice_number", ""),
        "date":            inv.get("date", ""),
        "customer_name":   inv.get("customer_name", ""),
        "gross_sale":      float(inv.get("subtotal",      0) or 0),
        "discount":        sum(float(it.get("discount", 0)) for it in inv.get("items", [])),
        "round_off":       float(inv.get("round_off",     0) or 0),
        "sale_amount":     float(inv.get("grand_total",   0) or 0),
        "cgst":            float(inv.get("cgst_amount",   0) or 0),
        "sgst":            float(inv.get("sgst_amount",   0) or 0),
        "other_purchase":  float(inv.get("old_purchase",  0) or 0),
        "advance":         float(inv.get("advance_paid",  0) or 0),
        "cash":            float(inv.get("cash_paid",     0) or 0),
        "card":            float(inv.get("card_paid",     0) or 0),
        "cheque":          float(inv.get("cheque_paid",   0) or 0),
        "dues":            float(inv.get("due_amount",    0) or 0),
        # filter helpers
        "_metals":     metals,
        "_purities":   purities,
        "_categories": categories,
    }


class SalesReportPage(QWidget):
    edit_invoice_requested      = pyqtSignal(dict)
    duplicate_invoice_requested = pyqtSignal(dict)

    def __init__(self, history_mode: bool = False):
        super().__init__()
        self.history_mode           = history_mode
        self._sort_col              = -1
        self._sort_asc              = True
        self._all_invoices: list    = []
        self._flat_rows: list       = []
        self._unsorted_flat: list   = []
        self._catalog_metal_map: dict = {}
        self._build_ui()

    # ─────────────────────────────────────────────────────────
    #  BUILD UI
    # ─────────────────────────────────────────────────────────
    def _build_ui(self):
        scroll = QScrollArea(self)
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(scroll)

        container = QWidget()
        scroll.setWidget(container)
        root = QVBoxLayout(container)
        root.setContentsMargins(25, 20, 25, 20)
        root.setSpacing(14)

        # Title
        page_title = "📋  Invoice History" if self.history_mode else "📊  Sales Report"
        title = QLabel(page_title)
        title.setFont(QFont("Segoe UI", 16, QFont.Bold))
        root.addWidget(title)

        # ── Quick date presets ────────────────────────────────
        preset_row = QHBoxLayout()
        preset_row.setSpacing(6)
        preset_row.addWidget(QLabel("Quick:"))
        _ps = (
            "QPushButton{background:#ecf0f1;color:#2c3e50;border:1px solid #bdc3c7;"
            "border-radius:4px;padding:4px 10px;font-size:11px;}"
            "QPushButton:hover{background:#d5d8dc;}"
        )
        for lbl, slot in [
            ("Today",      self._preset_today),
            ("This Week",  self._preset_week),
            ("This Month", self._preset_month),
            ("Last Month", self._preset_last_month),
            ("This Year",  self._preset_year),
            ("All Time",   self._preset_all),
        ]:
            b = QPushButton(lbl); b.setStyleSheet(_ps); b.clicked.connect(slot)
            preset_row.addWidget(b)
        preset_row.addStretch()
        root.addLayout(preset_row)

        # ── Filters ───────────────────────────────────────────
        filter_grp = QGroupBox("Filters")
        filter_grp.setStyleSheet("QGroupBox{font-weight:bold;}")
        fl = QVBoxLayout(filter_grp)
        fl.setSpacing(10)

        # Row 1 — dates, customer, invoice no
        row1 = QHBoxLayout(); row1.setSpacing(10)
        row1.addWidget(QLabel("From:"))
        self.dt_from = QDateEdit(QDate.currentDate().addMonths(-1))
        self.dt_from.setCalendarPopup(True); self.dt_from.setMinimumHeight(32)
        self.dt_from.dateChanged.connect(self._do_search)
        row1.addWidget(self.dt_from)

        row1.addWidget(QLabel("To:"))
        self.dt_to = QDateEdit(QDate.currentDate())
        self.dt_to.setCalendarPopup(True); self.dt_to.setMinimumHeight(32)
        self.dt_to.dateChanged.connect(self._do_search)
        row1.addWidget(self.dt_to)

        row1.addWidget(QLabel("Customer:"))
        self.txt_cust = QLineEdit()
        self.txt_cust.setPlaceholderText("Name or mobile")
        self.txt_cust.setMinimumHeight(32); self.txt_cust.setMaximumWidth(170)
        self.txt_cust.textChanged.connect(self._on_text_changed)
        row1.addWidget(self.txt_cust)

        row1.addWidget(QLabel("Invoice No:"))
        self.txt_inv = QLineEdit()
        self.txt_inv.setPlaceholderText("e.g. JB-0001")
        self.txt_inv.setMinimumHeight(32); self.txt_inv.setMaximumWidth(130)
        self.txt_inv.textChanged.connect(self._on_text_changed)
        row1.addWidget(self.txt_inv)

        btn_search = QPushButton("Search")
        btn_search.setStyleSheet(
            "QPushButton{background:#2980b9;color:white;border-radius:4px;padding:6px 14px;}"
            "QPushButton:hover{background:#2471a3;}")
        btn_search.setMinimumHeight(32); btn_search.clicked.connect(self._do_search)
        row1.addWidget(btn_search)

        btn_reset = QPushButton("Reset")
        btn_reset.setStyleSheet(
            "QPushButton{background:#7f8c8d;color:white;border-radius:4px;padding:6px 14px;}"
            "QPushButton:hover{background:#626567;}")
        btn_reset.setMinimumHeight(32); btn_reset.clicked.connect(self._reset_filters)
        row1.addWidget(btn_reset)
        row1.addStretch()
        fl.addLayout(row1)

        # Row 2 — category, metal, purity, due, min/max
        row2 = QHBoxLayout(); row2.setSpacing(10)

        def _cmb(items, min_w):
            c = QComboBox()
            c.addItems(items); c.setMinimumHeight(32); c.setMinimumWidth(min_w)
            c.currentIndexChanged.connect(self._do_search)
            return c

        row2.addWidget(QLabel("Category:"))
        self.cmb_category = _cmb(["All"], 120)
        row2.addWidget(self.cmb_category)

        row2.addWidget(QLabel("Metal:"))
        self.cmb_metal = _cmb(["All"], 100)
        row2.addWidget(self.cmb_metal)

        row2.addWidget(QLabel("Purity:"))
        self.cmb_purity = _cmb(["All"], 100)
        row2.addWidget(self.cmb_purity)

        row2.addWidget(QLabel("Due:"))
        self.cmb_due = _cmb(["All", "Paid (No Due)", "Has Due"], 120)
        row2.addWidget(self.cmb_due)

        row2.addWidget(QLabel("Min ₹:"))
        self.txt_min = QLineEdit()
        self.txt_min.setPlaceholderText("0")
        self.txt_min.setMinimumHeight(32); self.txt_min.setMaximumWidth(88)
        self.txt_min.textChanged.connect(self._on_text_changed)
        row2.addWidget(self.txt_min)

        row2.addWidget(QLabel("Max ₹:"))
        self.txt_max = QLineEdit()
        self.txt_max.setPlaceholderText("Any")
        self.txt_max.setMinimumHeight(32); self.txt_max.setMaximumWidth(88)
        self.txt_max.textChanged.connect(self._on_text_changed)
        row2.addWidget(self.txt_max)

        row2.addStretch()
        fl.addLayout(row2)
        root.addWidget(filter_grp)

        # ── Summary Cards ─────────────────────────────────────
        cards_row = QHBoxLayout(); cards_row.setSpacing(10)
        for card_title, default_val, color, attr in [
            ("Invoices",      "0",       "#2980b9", "lbl_count"),
            ("Gross Sale",    "₹ 0.00",  "#27ae60", "lbl_gross"),
            ("CGST + SGST",   "₹ 0.00",  "#f39c12", "lbl_gst"),
            ("Sale Amount",   "₹ 0.00",  "#8e44ad", "lbl_sale"),
            ("Total Dues",    "₹ 0.00",  "#e74c3c", "lbl_dues"),
        ]:
            frame = QFrame()
            frame.setFixedHeight(72)
            frame.setStyleSheet(_card_style(color))
            cl = QVBoxLayout(frame)
            cl.setContentsMargins(14, 8, 14, 8); cl.setSpacing(2)
            lt = QLabel(card_title)
            lt.setStyleSheet("color:rgba(255,255,255,0.85);font-size:11px;background:transparent;")
            lv = QLabel(default_val)
            lv.setFont(QFont("Segoe UI", 13, QFont.Bold))
            lv.setStyleSheet("color:white;background:transparent;")
            cl.addWidget(lt); cl.addWidget(lv)
            setattr(self, attr, lv)
            cards_row.addWidget(frame)
        root.addLayout(cards_row)

        # ── Payment breakdown bar ─────────────────────────────
        breakdown = QFrame()
        breakdown.setStyleSheet(
            "QFrame{background:#f8f9fa;border:1px solid #e0e0e0;border-radius:6px;}")
        bl = QHBoxLayout(breakdown)
        bl.setContentsMargins(14, 8, 14, 8); bl.setSpacing(16)
        hdr_lbl = QLabel("Payment Breakdown:")
        hdr_lbl.setFont(QFont("Segoe UI", 10, QFont.Bold))
        hdr_lbl.setStyleSheet("color:#2c3e50;background:transparent;")
        bl.addWidget(hdr_lbl)
        for mode, color, attr in [
            ("Cash",   "#27ae60", "lbl_cash"),
            ("Card",   "#2980b9", "lbl_card"),
            ("UPI",    "#8e44ad", "lbl_upi"),
            ("Cheque", "#f39c12", "lbl_cheque"),
        ]:
            pill = QLabel(f"{mode}: ₹ 0.00")
            pill.setStyleSheet(
                f"background:{color};color:white;border-radius:12px;"
                f"padding:3px 14px;font-size:11px;font-weight:bold;")
            setattr(self, attr, pill); bl.addWidget(pill)
        bl.addStretch()
        root.addWidget(breakdown)

        # ── Table ─────────────────────────────────────────────
        self.tbl = QTableWidget()
        self.tbl.setColumnCount(len(_RPT_LABELS))
        self.tbl.setHorizontalHeaderLabels(_RPT_LABELS)
        self.tbl.horizontalHeader().sectionClicked.connect(self._on_header_click)
        self.tbl.horizontalHeader().setCursor(Qt.PointingHandCursor)
        hdr = self.tbl.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.Stretch)
        self.tbl.verticalHeader().setDefaultSectionSize(38)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl.setAlternatingRowColors(True)
        self.tbl.setMinimumHeight(320)
        self.tbl.doubleClicked.connect(self._view_invoice)
        root.addWidget(self.tbl)

        # ── Action buttons ────────────────────────────────────
        act = QHBoxLayout()

        # Row-selection actions (left side, enabled when a row is selected)
        self._btn_edit_inv = QPushButton("✏️  Edit Invoice")
        self._btn_edit_inv.setEnabled(False)
        self._btn_edit_inv.setStyleSheet(
            "QPushButton{background:#2980b9;color:white;border-radius:4px;"
            "padding:8px 16px;font-weight:bold;}"
            "QPushButton:hover{background:#2471a3;}"
            "QPushButton:disabled{background:#bdc3c7;color:#7f8c8d;}"
        )
        self._btn_edit_inv.clicked.connect(self._edit_selected)

        self._btn_dup_inv = QPushButton("📋  Duplicate")
        self._btn_dup_inv.setEnabled(False)
        self._btn_dup_inv.setStyleSheet(
            "QPushButton{background:#8e44ad;color:white;border-radius:4px;"
            "padding:8px 16px;font-weight:bold;}"
            "QPushButton:hover{background:#7d3c98;}"
            "QPushButton:disabled{background:#bdc3c7;color:#7f8c8d;}"
        )
        self._btn_dup_inv.clicked.connect(self._duplicate_selected)

        act.addWidget(self._btn_edit_inv)
        act.addWidget(self._btn_dup_inv)
        act.addStretch()

        for label, color, hover, slot in [
            ("👁 Preview",   "#8e44ad", "#7d3c98", self._preview),
            ("Print / PDF",  "#f39c12", "#e67e22", self._reprint),
            ("Export Excel", "#1e8449", "#196f3d", self._export_excel),
            ("Export CSV",   "#2980b9", "#2471a3", self._export_csv),
        ]:
            b = QPushButton(label)
            b.setStyleSheet(
                f"QPushButton{{background:{color};color:white;border-radius:4px;padding:8px 16px;}}"
                f"QPushButton:hover{{background:{hover};}}")
            b.clicked.connect(slot); act.addWidget(b)
        root.addLayout(act)

        # Enable/disable row-selection buttons when selection changes
        self.tbl.itemSelectionChanged.connect(self._on_selection_changed)

    # ─────────────────────────────────────────────────────────
    #  Quick date presets
    # ─────────────────────────────────────────────────────────
    def _set_range(self, d_from: date, d_to: date):
        for w in (self.dt_from, self.dt_to): w.blockSignals(True)
        self.dt_from.setDate(QDate(d_from.year, d_from.month, d_from.day))
        self.dt_to.setDate(QDate(d_to.year, d_to.month, d_to.day))
        for w in (self.dt_from, self.dt_to): w.blockSignals(False)
        self._do_search()

    def _preset_today(self):
        t = date.today(); self._set_range(t, t)
    def _preset_week(self):
        t = date.today(); self._set_range(t - timedelta(days=t.weekday()), t)
    def _preset_month(self):
        t = date.today(); self._set_range(t.replace(day=1), t)
    def _preset_last_month(self):
        t = date.today(); end = t.replace(day=1) - timedelta(days=1)
        self._set_range(end.replace(day=1), end)
    def _preset_year(self):
        t = date.today(); self._set_range(t.replace(month=1, day=1), t)
    def _preset_all(self):
        for w in (self.dt_from, self.dt_to): w.blockSignals(True)
        self.dt_from.setDate(QDate(2000, 1, 1))
        self.dt_to.setDate(QDate.currentDate())
        for w in (self.dt_from, self.dt_to): w.blockSignals(False)
        self._do_search()

    # ─────────────────────────────────────────────────────────
    #  Refresh / dropdown loaders
    # ─────────────────────────────────────────────────────────
    def refresh(self):
        self._all_invoices = get_all_invoices()
        self._refresh_metals()
        self._refresh_categories()
        self._refresh_purities()
        self._do_search()

    def _refresh_metals(self):
        metals = get_metals()
        metal_id_map = {m['id']: m['name'] for m in metals if m.get('id')}
        self._catalog_metal_map = {}
        for ci in get_catalog():
            mid = ci.get('metal_id', '')
            if mid and mid in metal_id_map:
                self._catalog_metal_map[ci['name'].strip().lower()] = metal_id_map[mid]
        names = sorted({m['name'].strip() for m in metals if m.get('name')})
        self._reload_cmb(self.cmb_metal, names)

    def _refresh_categories(self):
        cats = set()
        for inv in self._all_invoices:
            for it in inv.get("items", []):
                c = it.get("category", "").strip()
                if c: cats.add(c)
        self._reload_cmb(self.cmb_category, sorted(cats))

    def _refresh_purities(self):
        purities = set()
        for inv in self._all_invoices:
            for it in inv.get("items", []):
                p = it.get("purity", "").strip()
                if p: purities.add(p)
        self._reload_cmb(self.cmb_purity, sorted(purities))

    def _reload_cmb(self, cmb: QComboBox, options: list):
        current = cmb.currentText()
        cmb.blockSignals(True)
        cmb.clear(); cmb.addItem("All")
        for o in options: cmb.addItem(o)
        idx = cmb.findText(current)
        cmb.setCurrentIndex(max(0, idx))
        cmb.blockSignals(False)

    def _reset_filters(self):
        for w in (self.dt_from, self.dt_to, self.cmb_category,
                  self.cmb_metal, self.cmb_purity, self.cmb_due):
            w.blockSignals(True)
        self.dt_from.setDate(QDate.currentDate().addMonths(-1))
        self.dt_to.setDate(QDate.currentDate())
        self.txt_cust.clear(); self.txt_inv.clear()
        self.txt_min.clear();  self.txt_max.clear()
        for c in (self.cmb_category, self.cmb_metal, self.cmb_purity, self.cmb_due):
            c.setCurrentIndex(0)
        for w in (self.dt_from, self.dt_to, self.cmb_category,
                  self.cmb_metal, self.cmb_purity, self.cmb_due):
            w.blockSignals(False)
        self._do_search()

    def _on_text_changed(self, text: str):
        if len(text.strip()) >= 3 or len(text.strip()) == 0:
            self._do_search()

    # ─────────────────────────────────────────────────────────
    #  Filter + build rows
    # ─────────────────────────────────────────────────────────
    def _do_search(self):
        d_from        = self.dt_from.date().toString("yyyy-MM-dd")
        d_to          = self.dt_to.date().toString("yyyy-MM-dd")
        cust_term     = self.txt_cust.text().strip().lower()
        inv_term      = self.txt_inv.text().strip().lower()
        cat_filter    = self.cmb_category.currentText()
        metal_filter  = self.cmb_metal.currentText().lower()
        purity_filter = self.cmb_purity.currentText().lower()
        due_filter    = self.cmb_due.currentText()

        try:    min_amt = float(self.txt_min.text()) if self.txt_min.text().strip() else None
        except: min_amt = None
        try:    max_amt = float(self.txt_max.text()) if self.txt_max.text().strip() else None
        except: max_amt = None

        rows = []
        for inv in self._all_invoices:
            if not (d_from <= inv.get("date", "") <= d_to):
                continue
            if cust_term and (
                cust_term not in inv.get("customer_name", "").lower() and
                cust_term not in inv.get("customer_mobile", "")
            ):
                continue
            if inv_term and inv_term not in inv.get("invoice_number", "").lower():
                continue

            gt = float(inv.get("grand_total", 0) or 0)
            if min_amt is not None and gt < min_amt: continue
            if max_amt is not None and gt > max_amt: continue

            due = _due_amount(inv)
            if due_filter == "Paid (No Due)" and due > 0: continue
            if due_filter == "Has Due"        and due <= 0: continue

            row = _build_inv_row(inv, self._catalog_metal_map)

            if cat_filter != "All" and cat_filter not in row["_categories"]:
                continue
            if metal_filter != "all" and metal_filter not in row["_metals"]:
                continue
            if purity_filter != "all" and purity_filter not in row["_purities"]:
                continue

            rows.append(row)

        self._unsorted_flat = rows
        self._flat_rows     = self._apply_sort(rows)
        self._populate(self._flat_rows)

    # ─────────────────────────────────────────────────────────
    #  Sort
    # ─────────────────────────────────────────────────────────
    def _on_header_click(self, col: int):
        if col in _NON_SORTABLE: return
        if self._sort_col == col:
            if self._sort_asc:
                self._sort_col = -1
                self._refresh_sort_headers()
                self._flat_rows = list(self._unsorted_flat)
                self._populate(self._flat_rows)
                return
            self._sort_asc = True
        else:
            self._sort_col = col
            self._sort_asc = False
        self._refresh_sort_headers()
        self._flat_rows = self._apply_sort(self._unsorted_flat)
        self._populate(self._flat_rows)

    def _apply_sort(self, rows: list) -> list:
        if self._sort_col < 0 or self._sort_col not in _SORT_KEYS:
            return list(rows)
        key = _SORT_KEYS[self._sort_col]
        def _k(r):
            v = r.get(key, "")
            try:    return (0, float(v))
            except: return (1, str(v).lower())
        return sorted(rows, key=_k, reverse=not self._sort_asc)

    def _refresh_sort_headers(self):
        for i, lbl in enumerate(_RPT_LABELS):
            item = self.tbl.horizontalHeaderItem(i)
            if not item: continue
            if i == self._sort_col:
                item.setText(lbl + (" ▲" if self._sort_asc else " ▼"))
            else:
                item.setText(lbl)

    # ─────────────────────────────────────────────────────────
    #  Populate table
    # ─────────────────────────────────────────────────────────
    def _populate(self, rows: list):
        self.tbl.setRowCount(0)
        self._btn_edit_inv.setEnabled(False)
        self._btn_dup_inv.setEnabled(False)
        t_gross = t_disc = t_sale = t_cgst = t_sgst = 0.0
        t_other = t_adv  = t_cash = t_card = 0.0
        t_cheque = t_dues = t_upi = 0.0

        for row in rows:
            r = self.tbl.rowCount()
            self.tbl.insertRow(r)
            inv = row["_inv"]

            vals = [
                row["invoice_number"],
                row["date"],
                row["customer_name"],
                f"{row['gross_sale']:,.2f}",
                f"{row['discount']:,.2f}",
                f"{row['sale_amount']:,.2f}",
                f"{row['cgst']:,.2f}",
                f"{row['sgst']:,.2f}",
                f"{row['other_purchase']:,.2f}",
                f"{row['advance']:,.2f}",
                f"{row['cash']:,.2f}",
                f"{row['card']:,.2f}",
                f"{row['cheque']:,.2f}",
                f"{row['dues']:,.2f}",
            ]
            _right = set(range(3, 14))
            for c, v in enumerate(vals):
                cell = QTableWidgetItem(v)
                cell.setTextAlignment(
                    Qt.AlignRight | Qt.AlignVCenter if c in _right
                    else Qt.AlignCenter
                )
                if c == 13 and row["dues"] > 0:
                    cell.setForeground(QColor("#e74c3c"))
                self.tbl.setItem(r, c, cell)

            t_gross  += row["gross_sale"]
            t_disc   += row["discount"]
            t_sale   += row["sale_amount"]
            t_cgst   += row["cgst"]
            t_sgst   += row["sgst"]
            t_other  += row["other_purchase"]
            t_adv    += row["advance"]
            t_cash   += row["cash"]
            t_card   += row["card"]
            t_cheque += row["cheque"]
            t_dues   += row["dues"]
            t_upi    += float(inv.get("upi_paid", 0) or 0)

        self.lbl_count.setText(str(len(rows)))
        self.lbl_gross.setText(format_currency(t_gross))
        self.lbl_gst.setText(format_currency(t_cgst + t_sgst))
        self.lbl_sale.setText(format_currency(t_sale))
        self.lbl_dues.setText(format_currency(t_dues))

        self.lbl_cash.setText(f"Cash: {format_currency(t_cash)}")
        self.lbl_card.setText(f"Card: {format_currency(t_card)}")
        self.lbl_upi.setText(f"UPI: {format_currency(t_upi)}")
        self.lbl_cheque.setText(f"Cheque: {format_currency(t_cheque)}")

    # ─────────────────────────────────────────────────────────
    #  Actions
    # ─────────────────────────────────────────────────────────
    def _open_detail(self, inv: dict):
        from ui.invoice_detail_dialog import InvoiceDetailDialog
        dlg = InvoiceDetailDialog(
            inv, self,
            on_edit=lambda i: self.edit_invoice_requested.emit(i),
            on_duplicate=lambda i: self.duplicate_invoice_requested.emit(i),
        )
        dlg.exec()

    def _view_invoice(self):
        row = self.tbl.currentRow()
        if 0 <= row < len(self._flat_rows):
            self._open_detail(self._flat_rows[row]["_inv"])

    def _on_selection_changed(self):
        row = self.tbl.currentRow()
        has_row = 0 <= row < len(self._flat_rows)
        self._btn_edit_inv.setEnabled(has_row)
        self._btn_dup_inv.setEnabled(has_row)

    def _edit_selected(self):
        row = self.tbl.currentRow()
        if 0 <= row < len(self._flat_rows):
            self.edit_invoice_requested.emit(self._flat_rows[row]["_inv"])

    def _duplicate_selected(self):
        row = self.tbl.currentRow()
        if 0 <= row < len(self._flat_rows):
            self.duplicate_invoice_requested.emit(self._flat_rows[row]["_inv"])

    def _preview(self):
        row = self.tbl.currentRow()
        if row < 0 or row >= len(self._flat_rows):
            QMessageBox.information(self, "Preview", "Select a row first.")
            return
        from ui.invoice_preview_dialog import InvoicePreviewDialog
        InvoicePreviewDialog(self._flat_rows[row]["_inv"], parent=self).exec()

    def _reprint(self):
        row = self.tbl.currentRow()
        if row < 0 or row >= len(self._flat_rows):
            QMessageBox.information(self, "Print", "Select a row first.")
            return
        save_invoice_as_pdf(self._flat_rows[row]["_inv"], parent=self)

    # ─────────────────────────────────────────────────────────
    #  Export CSV
    # ─────────────────────────────────────────────────────────
    def _export_csv(self):
        if not self._flat_rows:
            QMessageBox.information(self, "Export", "No data to export.")
            return
        from PyQt5.QtWidgets import QFileDialog
        path, _ = QFileDialog.getSaveFileName(
            self, "Save CSV", f"sales_report_{date.today()}.csv", "CSV Files (*.csv)")
        if not path: return
        try:
            fields = [
                "invoice_number", "date", "customer_name",
                "gross_sale", "discount", "sale_amount",
                "cgst", "sgst", "other_purchase",
                "advance", "cash", "card", "cheque", "dues",
            ]
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
                w.writeheader()
                for row in self._flat_rows:
                    w.writerow({k: row.get(k, "") for k in fields})
            QMessageBox.information(self, "Export", f"Exported to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    # ─────────────────────────────────────────────────────────
    #  Export Excel
    # ─────────────────────────────────────────────────────────
    def _export_excel(self):
        if not self._flat_rows:
            QMessageBox.information(self, "Export", "No data to export.")
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(self, "Missing Library",
                "openpyxl is not installed.\n\nRun: pip install openpyxl")
            return

        from PyQt5.QtWidgets import QFileDialog
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel", f"sales_report_{date.today()}.xlsx", "Excel Files (*.xlsx)")
        if not path: return

        try:
            wb   = openpyxl.Workbook()
            shop = AppConfig.shop()

            hdr_fill = PatternFill("solid", fgColor="1F4E79")
            hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
            ttl_font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
            sub_fill = PatternFill("solid", fgColor="D6E4F0")
            alt_fill = PatternFill("solid", fgColor="EBF5FB")
            thin     = Side(style="thin", color="BFBFBF")
            bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)
            ctr      = Alignment(horizontal="center", vertical="center")
            lft      = Alignment(horizontal="left",   vertical="center")
            mfmt     = '#,##0.00'

            def hc(ws, row, col, val):
                c = ws.cell(row=row, column=col, value=val)
                c.font=hdr_font; c.fill=hdr_fill; c.alignment=ctr; c.border=bdr
            def dc(ws, row, col, val, fmt=None, alt=False):
                c = ws.cell(row=row, column=col, value=val)
                c.alignment=ctr; c.border=bdr
                if alt: c.fill=alt_fill
                if fmt: c.number_format=fmt
            def aw(ws, mn=8, mx=40):
                for col in ws.columns:
                    w = mn
                    for cell in col:
                        try: w = max(w, len(str(cell.value or "")))
                        except: pass
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(w+2, mx)

            t_gross = sum(r["gross_sale"]  for r in self._flat_rows)
            t_disc  = sum(r["discount"]    for r in self._flat_rows)
            t_sale  = sum(r["sale_amount"] for r in self._flat_rows)
            t_cgst  = sum(r["cgst"]        for r in self._flat_rows)
            t_sgst  = sum(r["sgst"]        for r in self._flat_rows)
            t_other = sum(r["other_purchase"] for r in self._flat_rows)
            t_adv   = sum(r["advance"]     for r in self._flat_rows)
            t_cash  = sum(r["cash"]        for r in self._flat_rows)
            t_card  = sum(r["card"]        for r in self._flat_rows)
            t_cheq  = sum(r["cheque"]      for r in self._flat_rows)
            t_dues  = sum(r["dues"]        for r in self._flat_rows)
            t_upi   = sum(float(r["_inv"].get("upi_paid", 0) or 0) for r in self._flat_rows)

            # ════ Sheet 1 — Summary ═════════════════════════════
            ws1 = wb.active; ws1.title = "Summary"
            ws1.sheet_view.showGridLines = False
            ws1.merge_cells("A1:D1")
            t = ws1.cell(1, 1, f"{shop.get('shop_name','Jewelry Shop')}  —  Sales Report")
            t.font = ttl_font; t.alignment = lft; ws1.row_dimensions[1].height = 28
            ws1.merge_cells("A2:D2")
            p = ws1.cell(2, 1,
                f"Period: {self.dt_from.date().toString('yyyy-MM-dd')}  to  "
                f"{self.dt_to.date().toString('yyyy-MM-dd')}    |    Generated: {date.today()}")
            p.font = Font(name="Calibri", italic=True, color="7F7F7F", size=10)
            p.alignment = lft
            ws1.append([])
            for off, (lbl, val, fmt) in enumerate([
                ("Invoices",           len(self._flat_rows), None),
                ("Gross Sale (₹)",     t_gross,  mfmt),
                ("Discount (₹)",       t_disc,   mfmt),
                ("Sale Amount (₹)",    t_sale,   mfmt),
                ("CGST (₹)",           t_cgst,   mfmt),
                ("SGST (₹)",           t_sgst,   mfmt),
                ("Other Purchase (₹)", t_other,  mfmt),
                ("Advance (₹)",        t_adv,    mfmt),
                ("", "", None),
                ("Cash (₹)",           t_cash,   mfmt),
                ("Card (₹)",           t_card,   mfmt),
                ("UPI (₹)",            t_upi,    mfmt),
                ("Cheque (₹)",         t_cheq,   mfmt),
                ("Total Dues (₹)",     t_dues,   mfmt),
            ], start=4):
                ws1.row_dimensions[off].height = 22
                lc = ws1.cell(off, 1, lbl)
                lc.font = Font(name="Calibri", bold=bool(lbl), size=10)
                if lbl: lc.fill = sub_fill
                lc.alignment = lft; lc.border = bdr
                vc = ws1.cell(off, 2, val if lbl else "")
                vc.alignment = ctr; vc.border = bdr
                if fmt and lbl: vc.number_format = fmt
            aw(ws1)
            ws1.column_dimensions["A"].width = 26
            ws1.column_dimensions["B"].width = 18

            # ════ Sheet 2 — Invoice Details ══════════════════════
            ws2 = wb.create_sheet("Invoice Details")
            ws2.sheet_view.showGridLines = False
            ws2.freeze_panes = "A2"
            hdrs2 = [
                "Invoice No", "Date", "Customer", "Mobile",
                "Gross Sale (₹)", "Discount (₹)", "Sale Amt (₹)",
                "CGST (₹)", "SGST (₹)", "Other Purch (₹)",
                "Advance (₹)", "Cash (₹)", "Card (₹)",
                "UPI (₹)", "Cheque (₹)", "Dues (₹)", "Remarks",
            ]
            for c, h in enumerate(hdrs2, 1): hc(ws2, 1, c, h)
            ws2.row_dimensions[1].height = 24
            mn_c = set(range(5, 17))
            for r, row in enumerate(self._flat_rows, 2):
                inv = row["_inv"]
                alt = (r % 2 == 0)
                for c, v in enumerate([
                    row["invoice_number"],
                    row["date"],
                    row["customer_name"],
                    inv.get("customer_mobile", ""),
                    row["gross_sale"],
                    row["discount"],
                    row["sale_amount"],
                    row["cgst"],
                    row["sgst"],
                    row["other_purchase"],
                    row["advance"],
                    row["cash"],
                    row["card"],
                    float(inv.get("upi_paid", 0) or 0),
                    row["cheque"],
                    row["dues"],
                    inv.get("remarks", ""),
                ], 1):
                    dc(ws2, r, c, v, fmt=mfmt if c in mn_c else None, alt=alt)
                ws2.row_dimensions[r].height = 20
            aw(ws2)

            wb.save(path)
            QMessageBox.information(self, "Export Successful", f"Excel saved to:\n{path}")
            try: os.startfile(path)
            except: pass

        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))


# ============================================================
#  Item-wise Sales Report
#  Item-wise Sales Report
# ============================================================

_ITEM_LABELS = [
    "Date", "Invoice No", "Customer",
    "Item Name", "Qty", "Purity",
    "Gross (g)", "Less (g)", "Net (g)",
    "Rate ₹/g", "Making ₹", "Item Total ₹",
    "GST ₹", "Grand Total ₹",
]

_ITEM_SORT_KEYS = {
    0:  "date",           1:  "invoice_number",  2:  "customer_name",
    3:  "item_name",      4:  "qty",             5:  "purity",
    6:  "gross_wt",       7:  "less_wt",         8:  "net_wt",
    9:  "rate",           10: "making",          11: "item_total",
    12: "gst_amount",     13: "grand_total",
}

_ITEM_RIGHT = {4, 6, 7, 8, 9, 10, 11, 12, 13}   # right-align numeric columns


class ItemSalesReportPage(QWidget):
    edit_invoice_requested      = pyqtSignal(dict)
    duplicate_invoice_requested = pyqtSignal(dict)

    def __init__(self):
        super().__init__()
        self._all_invoices      = []
        self._catalog_metal_map = {}
        self._unsorted_rows     = []
        self._rows              = []
        self._sort_col          = -1
        self._sort_asc          = True
        self._build_ui()

    # ─────────────────────────────────────────────────────────
    #  UI
    # ─────────────────────────────────────────────────────────
    def _build_ui(self):
        scroll = QScrollArea(self)
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(scroll)

        container = QWidget()
        scroll.setWidget(container)
        root = QVBoxLayout(container)
        root.setContentsMargins(25, 20, 25, 20)
        root.setSpacing(14)

        # Title
        title = QLabel("📊  Sales Report  —  Item-wise")
        title.setFont(QFont("Segoe UI", 16, QFont.Bold))
        root.addWidget(title)

        # ── Quick presets ─────────────────────────────────────
        _ps = (
            "QPushButton{background:#ecf0f1;color:#2c3e50;border:1px solid #bdc3c7;"
            "border-radius:4px;padding:4px 10px;font-size:11px;}"
            "QPushButton:hover{background:#d5d8dc;}"
        )
        preset_row = QHBoxLayout(); preset_row.setSpacing(6)
        preset_row.addWidget(QLabel("Quick:"))
        for lbl, slot in [
            ("Today",      self._preset_today),
            ("Yesterday",  self._preset_yesterday),
            ("This Week",  self._preset_week),
            ("This Month", self._preset_month),
            ("Last Month", self._preset_last_month),
            ("This Year",  self._preset_year),
            ("All Time",   self._preset_all),
        ]:
            b = QPushButton(lbl); b.setStyleSheet(_ps); b.clicked.connect(slot)
            preset_row.addWidget(b)
        preset_row.addStretch()
        root.addLayout(preset_row)

        # ── Filters ───────────────────────────────────────────
        filter_grp = QGroupBox("Filters")
        filter_grp.setStyleSheet("QGroupBox{font-weight:bold;}")
        fl = QVBoxLayout(filter_grp); fl.setSpacing(8)

        def _le(ph, w=150):
            e = QLineEdit(); e.setPlaceholderText(ph)
            e.setMinimumHeight(30); e.setMaximumWidth(w)
            e.textChanged.connect(self._on_text_changed)
            return e

        def _cmb(w=120):
            c = QComboBox(); c.addItem("All")
            c.setMinimumHeight(30); c.setMinimumWidth(w)
            c.currentIndexChanged.connect(self._do_search)
            return c

        # Row 1 — dates + text filters
        r1 = QHBoxLayout(); r1.setSpacing(10)
        r1.addWidget(QLabel("From:"))
        self.dt_from = QDateEdit(QDate.currentDate())
        self.dt_from.setCalendarPopup(True); self.dt_from.setMinimumHeight(30)
        self.dt_from.dateChanged.connect(self._do_search)
        r1.addWidget(self.dt_from)

        r1.addWidget(QLabel("To:"))
        self.dt_to = QDateEdit(QDate.currentDate())
        self.dt_to.setCalendarPopup(True); self.dt_to.setMinimumHeight(30)
        self.dt_to.dateChanged.connect(self._do_search)
        r1.addWidget(self.dt_to)

        r1.addWidget(QLabel("Mobile:"))
        self.txt_mobile = _le("Customer mobile", 150)
        r1.addWidget(self.txt_mobile)

        r1.addWidget(QLabel("Item:"))
        self.txt_item = _le("Item name search", 170)
        r1.addWidget(self.txt_item)

        btn_srch = QPushButton("Search")
        btn_srch.setStyleSheet(
            "QPushButton{background:#2980b9;color:white;border-radius:4px;padding:6px 14px;}"
            "QPushButton:hover{background:#2471a3;}")
        btn_srch.setMinimumHeight(30); btn_srch.clicked.connect(self._do_search)
        r1.addWidget(btn_srch)

        btn_rst = QPushButton("Reset")
        btn_rst.setStyleSheet(
            "QPushButton{background:#7f8c8d;color:white;border-radius:4px;padding:6px 14px;}"
            "QPushButton:hover{background:#626567;}")
        btn_rst.setMinimumHeight(30); btn_rst.clicked.connect(self._reset_filters)
        r1.addWidget(btn_rst)
        r1.addStretch()
        fl.addLayout(r1)

        # Row 2 — metal, purity
        r2 = QHBoxLayout(); r2.setSpacing(10)
        r2.addWidget(QLabel("Metal:"))
        self.cmb_metal = _cmb(120); r2.addWidget(self.cmb_metal)
        r2.addWidget(QLabel("Purity:"))
        self.cmb_purity = _cmb(100); r2.addWidget(self.cmb_purity)
        r2.addStretch()
        fl.addLayout(r2)
        root.addWidget(filter_grp)

        # ── Summary cards ─────────────────────────────────────
        cards = QHBoxLayout(); cards.setSpacing(10)
        for ctitle, default, color, attr in [
            ("Items Sold",  "0",       "#2980b9", "lbl_count"),
            ("Net Weight",  "0.000 g", "#16a085", "lbl_net_wt"),
            ("Item Total",  "₹ 0.00",  "#27ae60", "lbl_sale"),
            ("Total GST",   "₹ 0.00",  "#f39c12", "lbl_gst"),
            ("Grand Total", "₹ 0.00",  "#8e44ad", "lbl_grand"),
        ]:
            fr = QFrame(); fr.setFixedHeight(72)
            fr.setStyleSheet(f"QFrame{{background:{color};border-radius:8px;border:none;}}")
            cl = QVBoxLayout(fr); cl.setContentsMargins(14, 8, 14, 8); cl.setSpacing(2)
            lt = QLabel(ctitle)
            lt.setStyleSheet("color:rgba(255,255,255,0.85);font-size:11px;background:transparent;")
            lv = QLabel(default)
            lv.setFont(QFont("Segoe UI", 13, QFont.Bold))
            lv.setStyleSheet("color:white;background:transparent;")
            cl.addWidget(lt); cl.addWidget(lv)
            setattr(self, attr, lv)
            cards.addWidget(fr)
        root.addLayout(cards)

        # ── Table ─────────────────────────────────────────────
        self.tbl = QTableWidget()
        self.tbl.setColumnCount(len(_ITEM_LABELS))
        self.tbl.setHorizontalHeaderLabels(_ITEM_LABELS)
        hdr = self.tbl.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(3, QHeaderView.Stretch)   # Item Name stretches
        hdr.sectionClicked.connect(self._on_header_click)
        hdr.setCursor(Qt.PointingHandCursor)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.verticalHeader().setDefaultSectionSize(36)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl.setAlternatingRowColors(True)
        self.tbl.setMinimumHeight(350)
        root.addWidget(self.tbl)

        # ── Totals footer ─────────────────────────────────────
        self._footer = QFrame()
        self._footer.setStyleSheet(
            "QFrame{background:#eaf4fb;border:1px solid #aed6f1;"
            "border-radius:5px;padding:2px;}")
        self._footer.setVisible(False)
        ftl = QHBoxLayout(self._footer)
        ftl.setContentsMargins(14, 6, 14, 6); ftl.setSpacing(20)
        for attr, label in [
            ("_ft_items",    "Items:"),
            ("_ft_net_wt",   "Net Wt:"),
            ("_ft_sale",     "Item Total:"),
            ("_ft_gst",      "GST:"),
            ("_ft_grand",    "Grand Total:"),
        ]:
            hl = QHBoxLayout(); hl.setSpacing(4)
            lbl = QLabel(label)
            lbl.setStyleSheet("font-weight:600;font-size:11px;color:#1a5276;background:transparent;")
            val = QLabel("—")
            val.setStyleSheet("font-weight:bold;font-size:12px;color:#154360;background:transparent;")
            hl.addWidget(lbl); hl.addWidget(val)
            setattr(self, attr, val)
            ftl.addLayout(hl)
        ftl.addStretch()
        root.addWidget(self._footer)

        # ── Export buttons ────────────────────────────────────
        act = QHBoxLayout(); act.addStretch()
        for label, color, hover, slot in [
            ("Export CSV",   "#2980b9", "#2471a3", self._export_csv),
            ("Export Excel", "#1e8449", "#196f3d", self._export_excel),
        ]:
            b = QPushButton(label)
            b.setStyleSheet(
                f"QPushButton{{background:{color};color:white;"
                f"border-radius:4px;padding:8px 16px;font-weight:bold;}}"
                f"QPushButton:hover{{background:{hover};}}")
            b.clicked.connect(slot); act.addWidget(b)
        root.addLayout(act)

    # ─────────────────────────────────────────────────────────
    #  Quick date presets
    # ─────────────────────────────────────────────────────────
    def _set_range(self, d_from, d_to):
        for w in (self.dt_from, self.dt_to): w.blockSignals(True)
        self.dt_from.setDate(QDate(d_from.year, d_from.month, d_from.day))
        self.dt_to.setDate(QDate(d_to.year,   d_to.month,   d_to.day))
        for w in (self.dt_from, self.dt_to): w.blockSignals(False)
        self._do_search()

    def _preset_today(self):
        t = date.today(); self._set_range(t, t)
    def _preset_yesterday(self):
        t = date.today() - timedelta(days=1); self._set_range(t, t)
    def _preset_week(self):
        t = date.today(); self._set_range(t - timedelta(days=t.weekday()), t)
    def _preset_month(self):
        t = date.today(); self._set_range(t.replace(day=1), t)
    def _preset_last_month(self):
        t = date.today(); end = t.replace(day=1) - timedelta(days=1)
        self._set_range(end.replace(day=1), end)
    def _preset_year(self):
        t = date.today(); self._set_range(t.replace(month=1, day=1), t)
    def _preset_all(self):
        for w in (self.dt_from, self.dt_to): w.blockSignals(True)
        self.dt_from.setDate(QDate(2000, 1, 1))
        self.dt_to.setDate(QDate.currentDate())
        for w in (self.dt_from, self.dt_to): w.blockSignals(False)
        self._do_search()

    # ─────────────────────────────────────────────────────────
    #  Refresh / dropdown loaders
    # ─────────────────────────────────────────────────────────
    def refresh(self):
        self._all_invoices = get_all_invoices()
        self._build_catalog_metal_map()
        self._refresh_dropdowns()
        self._do_search()

    def _build_catalog_metal_map(self):
        mid_map = {m["metal_id"]: m["name"] for m in get_metals() if m.get("metal_id")}
        self._catalog_metal_map = {}
        for ci in get_catalog():
            mid = ci.get("metal_id", "")
            if mid and mid in mid_map:
                self._catalog_metal_map[ci["name"].strip().lower()] = mid_map[mid]

    def _refresh_dropdowns(self):
        # Start with every metal defined in Settings so new metals appear immediately
        metals = {m["name"].strip() for m in get_metals() if m.get("name", "").strip()}
        purities = set()
        for inv in self._all_invoices:
            for it in inv.get("items", []):
                m = it.get("metal", "").strip()
                if not m:
                    m = self._catalog_metal_map.get(it.get("name", "").strip().lower(), "")
                p = it.get("purity", "").strip()
                if m: metals.add(m)
                if p: purities.add(p)
        self._reload_cmb(self.cmb_metal,  sorted(metals,   key=str.lower))
        self._reload_cmb(self.cmb_purity, sorted(purities, key=str.lower))

    def _reload_cmb(self, cmb, options):
        cur = cmb.currentText()
        cmb.blockSignals(True)
        cmb.clear(); cmb.addItem("All")
        for o in options: cmb.addItem(o)
        idx = cmb.findText(cur)
        cmb.setCurrentIndex(max(0, idx))
        cmb.blockSignals(False)

    def _reset_filters(self):
        for w in (self.dt_from, self.dt_to, self.cmb_metal, self.cmb_purity):
            w.blockSignals(True)
        self.dt_from.setDate(QDate.currentDate())
        self.dt_to.setDate(QDate.currentDate())
        self.txt_mobile.clear(); self.txt_item.clear()
        self.cmb_metal.setCurrentIndex(0); self.cmb_purity.setCurrentIndex(0)
        for w in (self.dt_from, self.dt_to, self.cmb_metal, self.cmb_purity):
            w.blockSignals(False)
        self._do_search()

    def _on_text_changed(self, text):
        if len(text.strip()) >= 2 or not text.strip():
            self._do_search()

    # ─────────────────────────────────────────────────────────
    #  Filter + build rows
    # ─────────────────────────────────────────────────────────
    def _do_search(self):
        d_from        = self.dt_from.date().toString("yyyy-MM-dd")
        d_to          = self.dt_to.date().toString("yyyy-MM-dd")
        mobile_term   = self.txt_mobile.text().strip()
        item_term     = self.txt_item.text().strip().lower()
        metal_filter  = self.cmb_metal.currentText()
        purity_filter = self.cmb_purity.currentText()

        rows = []
        for inv in self._all_invoices:
            inv_date = inv.get("date", "")
            if not (d_from <= inv_date <= d_to):
                continue
            if mobile_term and mobile_term not in inv.get("customer_mobile", ""):
                continue

            subtotal   = float(inv.get("subtotal",    0) or 0)
            tax_amount = float(inv.get("tax_amount",  0) or 0)

            for it in inv.get("items", []):
                name = it.get("name", "").strip()
                if not name:
                    continue

                # item name filter
                if item_term and item_term not in name.lower():
                    continue

                # resolve metal (fall back to catalog map if blank on item)
                it_metal = it.get("metal", "").strip()
                if not it_metal:
                    it_metal = self._catalog_metal_map.get(name.lower(), "")

                it_purity = it.get("purity", "").strip()

                # metal filter
                if metal_filter != "All" and it_metal.lower() != metal_filter.lower():
                    continue
                # purity filter
                if purity_filter != "All" and it_purity.lower() != purity_filter.lower():
                    continue

                item_total = float(it.get("total", 0) or 0)
                # distribute invoice GST proportionally across items by their value share
                gst_amount = (
                    round(item_total / subtotal * tax_amount, 2)
                    if subtotal > 0 else 0.0
                )

                rows.append({
                    "_inv":            inv,
                    "date":            inv_date,
                    "invoice_number":  inv.get("invoice_number", ""),
                    "customer_name":   inv.get("customer_name",  ""),
                    "customer_mobile": inv.get("customer_mobile",""),
                    "item_name":       name,
                    "qty":             int(it.get("quantity",     1) or 1),
                    "purity":          it_purity,
                    "gross_wt":        float(it.get("weight",       0) or 0),
                    "less_wt":         float(it.get("less_weight",  0) or 0),
                    "net_wt":          float(it.get("nett_weight",  0) or 0),
                    "rate":            float(it.get("rate",         0) or 0),
                    "making":          float(it.get("making_charge",0) or 0),
                    "item_total":      item_total,
                    "gst_amount":      gst_amount,
                    "grand_total":     round(item_total + gst_amount, 2),
                })

        self._unsorted_rows = rows
        self._rows          = self._apply_sort(rows)
        self._populate(self._rows)

    # ─────────────────────────────────────────────────────────
    #  Sort
    # ─────────────────────────────────────────────────────────
    def _on_header_click(self, col):
        if self._sort_col == col:
            if self._sort_asc:           # asc → remove sort
                self._sort_col = -1
                self._refresh_sort_headers()
                self._rows = list(self._unsorted_rows)
                self._populate(self._rows)
                return
            self._sort_asc = True        # desc → asc
        else:
            self._sort_col = col
            self._sort_asc = False       # first click → desc
        self._refresh_sort_headers()
        self._rows = self._apply_sort(self._unsorted_rows)
        self._populate(self._rows)

    def _apply_sort(self, rows):
        if self._sort_col < 0 or self._sort_col not in _ITEM_SORT_KEYS:
            return list(rows)
        key = _ITEM_SORT_KEYS[self._sort_col]
        def _k(r):
            v = r.get(key, "")
            try:    return (0, float(v))
            except: return (1, str(v).lower())
        return sorted(rows, key=_k, reverse=not self._sort_asc)

    def _refresh_sort_headers(self):
        for i, lbl in enumerate(_ITEM_LABELS):
            item = self.tbl.horizontalHeaderItem(i)
            if not item: continue
            if i == self._sort_col:
                item.setText(lbl + (" ▲" if self._sort_asc else " ▼"))
            else:
                item.setText(lbl)

    # ─────────────────────────────────────────────────────────
    #  Populate table
    # ─────────────────────────────────────────────────────────
    def _populate(self, rows):
        self.tbl.setRowCount(0)

        t_net = 0.0; t_sale = 0.0; t_gst = 0.0; t_grand = 0.0; t_qty = 0

        for row in rows:
            r = self.tbl.rowCount()
            self.tbl.insertRow(r)

            cells = [
                row["date"],
                row["invoice_number"],
                row["customer_name"],
                row["item_name"],
                str(row["qty"]),
                row["purity"],
                f"{row['gross_wt']:.3f}",
                f"{row['less_wt']:.3f}",
                f"{row['net_wt']:.3f}",
                f"{row['rate']:,.2f}",
                f"{row['making']:,.2f}",
                f"{row['item_total']:,.2f}",
                f"{row['gst_amount']:,.2f}",
                f"{row['grand_total']:,.2f}",
            ]

            for c, v in enumerate(cells):
                cell = QTableWidgetItem(v)
                align = Qt.AlignRight | Qt.AlignVCenter if c in _ITEM_RIGHT else Qt.AlignCenter
                cell.setTextAlignment(align)
                self.tbl.setItem(r, c, cell)

            t_qty   += row["qty"]
            t_net   += row["net_wt"]
            t_sale  += row["item_total"]
            t_gst   += row["gst_amount"]
            t_grand += row["grand_total"]

        # Summary cards
        self.lbl_count.setText(str(len(rows)))
        self.lbl_net_wt.setText(f"{t_net:.3f} g")
        self.lbl_sale.setText(format_currency(t_sale))
        self.lbl_gst.setText(format_currency(t_gst))
        self.lbl_grand.setText(format_currency(t_grand))

        # Footer bar
        if rows:
            self._ft_items.setText(str(len(rows)))
            self._ft_net_wt.setText(f"{t_net:.3f} g")
            self._ft_sale.setText(format_currency(t_sale))
            self._ft_gst.setText(format_currency(t_gst))
            self._ft_grand.setText(format_currency(t_grand))
            self._footer.setVisible(True)
        else:
            self._footer.setVisible(False)

    # ─────────────────────────────────────────────────────────
    #  Export CSV
    # ─────────────────────────────────────────────────────────
    def _export_csv(self):
        if not self._rows:
            QMessageBox.information(self, "Export", "No data to export."); return
        from PyQt5.QtWidgets import QFileDialog
        path, _ = QFileDialog.getSaveFileName(
            self, "Save CSV", f"sales_items_{date.today()}.csv", "CSV Files (*.csv)")
        if not path: return
        try:
            fields = [
                "date", "invoice_number", "customer_name", "customer_mobile",
                "item_name", "qty", "purity",
                "gross_wt", "less_wt", "net_wt",
                "rate", "making", "item_total", "gst_amount", "grand_total",
            ]
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
                w.writeheader()
                for row in self._rows:
                    w.writerow({k: row.get(k, "") for k in fields})
            QMessageBox.information(self, "Export", f"Exported to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    # ─────────────────────────────────────────────────────────
    #  Export Excel
    # ─────────────────────────────────────────────────────────
    def _export_excel(self):
        if not self._rows:
            QMessageBox.information(self, "Export", "No data to export."); return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(self, "Missing Library",
                "openpyxl is not installed.\n\nRun: pip install openpyxl")
            return

        from PyQt5.QtWidgets import QFileDialog
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel", f"sales_items_{date.today()}.xlsx", "Excel Files (*.xlsx)")
        if not path: return

        try:
            wb  = openpyxl.Workbook()
            ws  = wb.active; ws.title = "Item Sales"
            ws.sheet_view.showGridLines = False
            ws.freeze_panes = "A2"

            thin      = Side(style="thin", color="BFBFBF")
            bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
            ctr       = Alignment(horizontal="center", vertical="center")
            lft       = Alignment(horizontal="left",   vertical="center")
            hdr_fill  = PatternFill("solid", fgColor="1F4E79")
            hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
            alt_fill  = PatternFill("solid", fgColor="EBF5FB")
            tot_fill  = PatternFill("solid", fgColor="D6E4F0")
            tot_font  = Font(name="Calibri", bold=True, size=11)
            mfmt      = "#,##0.00"
            wfmt      = "#,##0.000"

            xlsx_hdrs = [
                "Date", "Invoice No", "Customer", "Mobile",
                "Item Name", "Qty", "Purity",
                "Gross (g)", "Less (g)", "Net (g)",
                "Rate ₹/g", "Making ₹", "Item Total ₹", "GST ₹", "Grand Total ₹",
            ]
            _money_c = {11, 12, 13, 14, 15}
            _wt_c    = {8, 9, 10}

            for c, h in enumerate(xlsx_hdrs, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = hdr_font; cell.fill = hdr_fill
                cell.alignment = ctr; cell.border = bdr
            ws.row_dimensions[1].height = 24

            for ri, row in enumerate(self._rows, 2):
                inv = row["_inv"]
                alt = (ri % 2 == 0)
                data = [
                    row["date"],            row["invoice_number"],
                    row["customer_name"],   inv.get("customer_mobile", ""),
                    row["item_name"],       row["qty"],
                    row["purity"],          row["gross_wt"],
                    row["less_wt"],         row["net_wt"],
                    row["rate"],            row["making"],
                    row["item_total"],      row["gst_amount"],
                    row["grand_total"],
                ]
                for c, v in enumerate(data, 1):
                    cell = ws.cell(row=ri, column=c, value=v)
                    cell.alignment = ctr; cell.border = bdr
                    if alt: cell.fill = alt_fill
                    if   c in _wt_c:    cell.number_format = wfmt
                    elif c in _money_c: cell.number_format = mfmt
                ws.row_dimensions[ri].height = 20

            tr = len(self._rows) + 2
            ws.row_dimensions[tr].height = 22
            for c in range(1, len(xlsx_hdrs) + 1):
                cell = ws.cell(row=tr, column=c)
                cell.fill = tot_fill; cell.border = bdr
                cell.alignment = ctr; cell.font = tot_font
            ws.cell(row=tr, column=1, value="TOTALS")

            totals_map = {
                6:  sum(r["qty"]         for r in self._rows),
                8:  sum(r["gross_wt"]    for r in self._rows),
                9:  sum(r["less_wt"]     for r in self._rows),
                10: sum(r["net_wt"]      for r in self._rows),
                13: sum(r["item_total"]  for r in self._rows),
                14: sum(r["gst_amount"]  for r in self._rows),
                15: sum(r["grand_total"] for r in self._rows),
            }
            for c, v in totals_map.items():
                cell = ws.cell(row=tr, column=c, value=v)
                cell.fill = tot_fill; cell.border = bdr
                cell.alignment = ctr; cell.font = tot_font
                if   c in _wt_c:    cell.number_format = wfmt
                elif c in _money_c: cell.number_format = mfmt

            for col in ws.columns:
                w = 8
                for cell in col:
                    try: w = max(w, len(str(cell.value or "")))
                    except: pass
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 40)

            shop = AppConfig.shop()
            ws2 = wb.create_sheet("Summary")
            ws2.sheet_view.showGridLines = False
            ws2.merge_cells("A1:B1")
            t = ws2.cell(1, 1, f"{shop.get('shop_name','Jewelry Shop')}  —  Item-wise Sales")
            t.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
            t.alignment = lft
            ws2.merge_cells("A2:B2")
            p = ws2.cell(2, 1,
                f"Period: {self.dt_from.date().toString('yyyy-MM-dd')}  to  "
                f"{self.dt_to.date().toString('yyyy-MM-dd')}    |    Generated: {date.today()}")
            p.font = Font(name="Calibri", italic=True, color="7F7F7F", size=10)
            p.alignment = lft
            ws2.append([])
            sub_fill = PatternFill("solid", fgColor="D6E4F0")
            for row_label, val, fmt in [
                ("Items Sold",       len(self._rows),                             None),
                ("Total Net Wt (g)", sum(r["net_wt"]     for r in self._rows),   wfmt),
                ("Item Total (₹)",   sum(r["item_total"] for r in self._rows),   mfmt),
                ("Total GST (₹)",    sum(r["gst_amount"] for r in self._rows),   mfmt),
                ("Grand Total (₹)",  sum(r["grand_total"]for r in self._rows),   mfmt),
            ]:
                r_idx = ws2.max_row + 1
                lc = ws2.cell(r_idx, 1, row_label)
                lc.font = Font(name="Calibri", bold=True, size=10)
                lc.fill = sub_fill; lc.alignment = lft; lc.border = bdr
                vc = ws2.cell(r_idx, 2, val)
                vc.alignment = ctr; vc.border = bdr
                if fmt: vc.number_format = fmt
            ws2.column_dimensions["A"].width = 24
            ws2.column_dimensions["B"].width = 18

            wb.save(path)
            QMessageBox.information(self, "Export Successful", f"Excel saved to:\n{path}")
            try: os.startfile(path)
            except: pass

        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))


# ============================================================
#  Date-wise Sales Report  (one row per invoice)
# ============================================================

_DWISE_LABELS = [
    "Date", "Invoice No.", "Customer",
    "Gross Sale ₹", "Discount ₹", "Sale Amt ₹",
    "CGST ₹", "SGST ₹", "IGST ₹",
    "Advance ₹", "Cash ₹", "Card ₹", "UPI ₹", "Cheque ₹", "Dues ₹",
]
_DWISE_RIGHT = {3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14}
_DWISE_SORT_KEYS = {
    0:  "date",           1:  "invoice_number",  2:  "customer_name",
    3:  "gross_sale",     4:  "discount",        5:  "sale_amt",
    6:  "cgst",           7:  "sgst",            8:  "igst",
    9:  "advance",        10: "cash",            11: "card",
    12: "upi",            13: "cheque",          14: "dues",
}


class DatewiseSalesPage(QWidget):

    def __init__(self):
        super().__init__()
        self._all_invoices  = []
        self._unsorted_rows = []
        self._rows          = []
        self._sort_col      = -1
        self._sort_asc      = True
        self._build_ui()

    # ─────────────────────────────────────────────────────────
    #  UI
    # ─────────────────────────────────────────────────────────
    def _build_ui(self):
        scroll = QScrollArea(self)
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(scroll)

        container = QWidget()
        scroll.setWidget(container)
        root = QVBoxLayout(container)
        root.setContentsMargins(25, 20, 25, 20)
        root.setSpacing(14)

        title = QLabel("\U0001f4c5  Sales Report  —  Date-wise")
        title.setFont(QFont("Segoe UI", 16, QFont.Bold))
        root.addWidget(title)

        _ps = (
            "QPushButton{background:#ecf0f1;color:#2c3e50;border:1px solid #bdc3c7;"
            "border-radius:4px;padding:4px 10px;font-size:11px;}"
            "QPushButton:hover{background:#d5d8dc;}"
        )
        preset_row = QHBoxLayout(); preset_row.setSpacing(6)
        preset_row.addWidget(QLabel("Quick:"))
        for lbl, slot in [
            ("Today",      self._preset_today),
            ("Yesterday",  self._preset_yesterday),
            ("This Week",  self._preset_week),
            ("This Month", self._preset_month),
            ("Last Month", self._preset_last_month),
            ("This Year",  self._preset_year),
            ("All Time",   self._preset_all),
        ]:
            b = QPushButton(lbl); b.setStyleSheet(_ps); b.clicked.connect(slot)
            preset_row.addWidget(b)
        preset_row.addStretch()
        root.addLayout(preset_row)

        filter_grp = QGroupBox("Filters")
        filter_grp.setStyleSheet("QGroupBox{font-weight:bold;}")
        fl = QHBoxLayout(filter_grp); fl.setSpacing(10)
        fl.addWidget(QLabel("From:"))
        self.dt_from = QDateEdit(QDate.currentDate())
        self.dt_from.setCalendarPopup(True); self.dt_from.setMinimumHeight(30)
        self.dt_from.dateChanged.connect(self._do_search)
        fl.addWidget(self.dt_from)
        fl.addWidget(QLabel("To:"))
        self.dt_to = QDateEdit(QDate.currentDate())
        self.dt_to.setCalendarPopup(True); self.dt_to.setMinimumHeight(30)
        self.dt_to.dateChanged.connect(self._do_search)
        fl.addWidget(self.dt_to)
        btn_rst = QPushButton("Reset")
        btn_rst.setStyleSheet(
            "QPushButton{background:#7f8c8d;color:white;border-radius:4px;padding:6px 14px;}"
            "QPushButton:hover{background:#626567;}")
        btn_rst.setMinimumHeight(30); btn_rst.clicked.connect(self._reset_filters)
        fl.addWidget(btn_rst)
        fl.addStretch()
        root.addWidget(filter_grp)

        cards = QHBoxLayout(); cards.setSpacing(10)
        for ctitle, default, color, attr in [
            ("Invoices",            "0",       "#2980b9", "lbl_inv_count"),
            ("Gross Sale",          "₹ 0.00",  "#27ae60", "lbl_gross"),
            ("Total GST",           "₹ 0.00",  "#f39c12", "lbl_gst"),
            ("Sale Amt",            "₹ 0.00",  "#8e44ad", "lbl_sale"),
            ("Revenue Collected",   "₹ 0.00",  "#1e8449", "lbl_revenue"),
            ("Total Dues",          "₹ 0.00",  "#e74c3c", "lbl_dues"),
        ]:
            fr = QFrame(); fr.setFixedHeight(72)
            fr.setStyleSheet(f"QFrame{{background:{color};border-radius:8px;border:none;}}")
            cl = QVBoxLayout(fr); cl.setContentsMargins(14, 8, 14, 8); cl.setSpacing(2)
            lt = QLabel(ctitle)
            lt.setStyleSheet("color:rgba(255,255,255,0.85);font-size:11px;background:transparent;")
            lv = QLabel(default)
            lv.setFont(QFont("Segoe UI", 13, QFont.Bold))
            lv.setStyleSheet("color:white;background:transparent;")
            cl.addWidget(lt); cl.addWidget(lv)
            setattr(self, attr, lv)
            cards.addWidget(fr)
        root.addLayout(cards)

        self.tbl = QTableWidget()
        self.tbl.setColumnCount(len(_DWISE_LABELS))
        self.tbl.setHorizontalHeaderLabels(_DWISE_LABELS)
        hdr = self.tbl.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(2, QHeaderView.Stretch)
        hdr.sectionClicked.connect(self._on_header_click)
        hdr.setCursor(Qt.PointingHandCursor)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.verticalHeader().setDefaultSectionSize(36)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl.setAlternatingRowColors(True)
        self.tbl.setMinimumHeight(350)
        root.addWidget(self.tbl)

        self._footer = QFrame()
        self._footer.setStyleSheet(
            "QFrame{background:#eaf4fb;border:1px solid #aed6f1;"
            "border-radius:5px;padding:2px;}")
        self._footer.setVisible(False)
        ftl = QHBoxLayout(self._footer)
        ftl.setContentsMargins(14, 6, 14, 6); ftl.setSpacing(20)
        for attr, label in [
            ("_ft_inv",      "Invoices:"),
            ("_ft_gross",    "Gross Sale:"),
            ("_ft_discount", "Discount:"),
            ("_ft_sale",     "Sale Amt:"),
            ("_ft_gst",      "GST:"),
            ("_ft_revenue",  "Revenue:"),
            ("_ft_dues",     "Dues:"),
            ("_ft_cash",     "Cash:"),
            ("_ft_card",     "Card:"),
            ("_ft_upi",      "UPI:"),
            ("_ft_cheque",   "Cheque:"),
        ]:
            hl = QHBoxLayout(); hl.setSpacing(4)
            lbl = QLabel(label)
            lbl.setStyleSheet("font-weight:600;font-size:11px;color:#1a5276;background:transparent;")
            val = QLabel("—")
            val.setStyleSheet("font-weight:bold;font-size:12px;color:#154360;background:transparent;")
            hl.addWidget(lbl); hl.addWidget(val)
            setattr(self, attr, val)
            ftl.addLayout(hl)
        ftl.addStretch()
        root.addWidget(self._footer)

        act = QHBoxLayout(); act.addStretch()
        for label, color, hover, slot in [
            ("Export CSV",   "#2980b9", "#2471a3", self._export_csv),
            ("Export Excel", "#1e8449", "#196f3d", self._export_excel),
        ]:
            b = QPushButton(label)
            b.setStyleSheet(
                f"QPushButton{{background:{color};color:white;"
                f"border-radius:4px;padding:8px 16px;font-weight:bold;}}"
                f"QPushButton:hover{{background:{hover};}}")
            b.clicked.connect(slot); act.addWidget(b)
        root.addLayout(act)

    # ─────────────────────────────────────────────────────────
    #  Quick date presets
    # ─────────────────────────────────────────────────────────
    def _set_range(self, d_from, d_to):
        for w in (self.dt_from, self.dt_to): w.blockSignals(True)
        self.dt_from.setDate(QDate(d_from.year, d_from.month, d_from.day))
        self.dt_to.setDate(QDate(d_to.year,     d_to.month,   d_to.day))
        for w in (self.dt_from, self.dt_to): w.blockSignals(False)
        self._do_search()

    def _preset_today(self):
        t = date.today(); self._set_range(t, t)
    def _preset_yesterday(self):
        t = date.today() - timedelta(days=1); self._set_range(t, t)
    def _preset_week(self):
        t = date.today(); self._set_range(t - timedelta(days=t.weekday()), t)
    def _preset_month(self):
        t = date.today(); self._set_range(t.replace(day=1), t)
    def _preset_last_month(self):
        t = date.today(); end = t.replace(day=1) - timedelta(days=1)
        self._set_range(end.replace(day=1), end)
    def _preset_year(self):
        t = date.today(); self._set_range(t.replace(month=1, day=1), t)
    def _preset_all(self):
        for w in (self.dt_from, self.dt_to): w.blockSignals(True)
        self.dt_from.setDate(QDate(2000, 1, 1))
        self.dt_to.setDate(QDate.currentDate())
        for w in (self.dt_from, self.dt_to): w.blockSignals(False)
        self._do_search()

    def _reset_filters(self):
        for w in (self.dt_from, self.dt_to): w.blockSignals(True)
        self.dt_from.setDate(QDate.currentDate())
        self.dt_to.setDate(QDate.currentDate())
        for w in (self.dt_from, self.dt_to): w.blockSignals(False)
        self._do_search()

    # ─────────────────────────────────────────────────────────
    #  Refresh / search
    # ─────────────────────────────────────────────────────────
    def refresh(self):
        self._all_invoices = get_all_invoices()
        self._do_search()

    def _do_search(self):
        d_from = self.dt_from.date().toString("yyyy-MM-dd")
        d_to   = self.dt_to.date().toString("yyyy-MM-dd")
        rows = []
        for inv in self._all_invoices:
            inv_date = inv.get("date", "")
            if not inv_date or not (d_from <= inv_date <= d_to):
                continue
            discount = sum(float(it.get("discount", 0) or 0) for it in inv.get("items", []))
            rows.append({
                "_inv":           inv,
                "date":           inv_date,
                "invoice_number": inv.get("invoice_number", ""),
                "customer_name":  inv.get("customer_name",  ""),
                "gross_sale":     float(inv.get("subtotal",     0) or 0),
                "discount":       discount,
                "sale_amt":       float(inv.get("grand_total",  0) or 0),
                "cgst":           float(inv.get("cgst_amount",  0) or 0),
                "sgst":           float(inv.get("sgst_amount",  0) or 0),
                "igst":           float(inv.get("igst_amount",  0) or 0),
                "advance":        float(inv.get("advance_paid", 0) or 0),
                "cash":           float(inv.get("cash_paid",    0) or 0),
                "card":           float(inv.get("card_paid",    0) or 0),
                "upi":            float(inv.get("upi_paid",     0) or 0),
                "cheque":         float(inv.get("cheque_paid",  0) or 0),
                "dues":           float(inv.get("due_amount",   0) or 0),
            })
        self._unsorted_rows = rows
        self._rows = self._apply_sort(rows)
        self._populate(self._rows)

    # ─────────────────────────────────────────────────────────
    #  Sort
    # ─────────────────────────────────────────────────────────
    def _on_header_click(self, col):
        if self._sort_col == col:
            if self._sort_asc:
                self._sort_col = -1
                self._refresh_sort_headers()
                self._rows = self._apply_sort(self._unsorted_rows)
                self._populate(self._rows)
                return
            self._sort_asc = True
        else:
            self._sort_col = col
            self._sort_asc = False
        self._refresh_sort_headers()
        self._rows = self._apply_sort(self._unsorted_rows)
        self._populate(self._rows)

    def _apply_sort(self, rows):
        if self._sort_col < 0 or self._sort_col not in _DWISE_SORT_KEYS:
            return sorted(rows, key=lambda r: (r["date"], r["invoice_number"]), reverse=True)
        key = _DWISE_SORT_KEYS[self._sort_col]
        def _k(r):
            v = r.get(key, "")
            try:    return (0, float(v))
            except: return (1, str(v).lower())
        return sorted(rows, key=_k, reverse=not self._sort_asc)

    def _refresh_sort_headers(self):
        for i, lbl in enumerate(_DWISE_LABELS):
            item = self.tbl.horizontalHeaderItem(i)
            if not item: continue
            if i == self._sort_col:
                item.setText(lbl + (" ▲" if self._sort_asc else " ▼"))
            else:
                item.setText(lbl)

    # ─────────────────────────────────────────────────────────
    #  Populate table
    # ─────────────────────────────────────────────────────────
    def _populate(self, rows):
        self.tbl.setRowCount(0)
        t_gross = 0.0; t_disc  = 0.0; t_sale = 0.0
        t_cgst  = 0.0; t_sgst  = 0.0; t_igst = 0.0
        t_cash  = 0.0; t_card  = 0.0; t_upi  = 0.0
        t_cheq  = 0.0; t_dues  = 0.0; t_adv  = 0.0

        for row in rows:
            r = self.tbl.rowCount(); self.tbl.insertRow(r)
            dues = row["dues"]
            cells = [
                row["date"],
                row["invoice_number"],
                row["customer_name"],
                f"{row['gross_sale']:,.2f}",
                f"{row['discount']:,.2f}",
                f"{row['sale_amt']:,.2f}",
                f"{row['cgst']:,.2f}",
                f"{row['sgst']:,.2f}",
                f"{row['igst']:,.2f}",
                f"{row['advance']:,.2f}",
                f"{row['cash']:,.2f}",
                f"{row['card']:,.2f}",
                f"{row['upi']:,.2f}",
                f"{row['cheque']:,.2f}",
                f"{dues:,.2f}",
            ]
            for c, v in enumerate(cells):
                cell = QTableWidgetItem(v)
                align = Qt.AlignRight | Qt.AlignVCenter if c in _DWISE_RIGHT else Qt.AlignCenter
                cell.setTextAlignment(align)
                if c == 14 and dues > 0:
                    cell.setForeground(QColor("#c0392b"))
                    f = cell.font(); f.setBold(True); cell.setFont(f)
                self.tbl.setItem(r, c, cell)

            t_gross += row["gross_sale"]
            t_disc  += row["discount"]
            t_sale  += row["sale_amt"]
            t_cgst  += row["cgst"]
            t_sgst  += row["sgst"]
            t_igst  += row["igst"]
            t_cash  += row["cash"]
            t_card  += row["card"]
            t_upi   += row["upi"]
            t_cheq  += row["cheque"]
            t_dues  += row["dues"]
            t_adv   += row["advance"]

        t_gst      = t_cgst + t_sgst + t_igst
        t_revenue  = round(t_sale - t_dues, 2)

        self.lbl_inv_count.setText(str(len(rows)))
        self.lbl_gross.setText(format_currency(t_gross))
        self.lbl_gst.setText(format_currency(t_gst))
        self.lbl_sale.setText(format_currency(t_sale))
        self.lbl_revenue.setText(format_currency(t_revenue))
        self.lbl_dues.setText(format_currency(t_dues))

        if rows:
            self._ft_inv.setText(str(len(rows)))
            self._ft_gross.setText(format_currency(t_gross))
            self._ft_discount.setText(format_currency(t_disc))
            self._ft_sale.setText(format_currency(t_sale))
            self._ft_gst.setText(format_currency(t_gst))
            self._ft_revenue.setText(format_currency(t_revenue))
            self._ft_dues.setText(format_currency(t_dues))
            self._ft_cash.setText(format_currency(t_cash))
            self._ft_card.setText(format_currency(t_card))
            self._ft_upi.setText(format_currency(t_upi))
            self._ft_cheque.setText(format_currency(t_cheq))
            self._footer.setVisible(True)
        else:
            self._footer.setVisible(False)

    # ─────────────────────────────────────────────────────────
    #  Export CSV
    # ─────────────────────────────────────────────────────────
    def _export_csv(self):
        if not self._rows:
            QMessageBox.information(self, "Export", "No data to export."); return
        from PyQt5.QtWidgets import QFileDialog
        path, _ = QFileDialog.getSaveFileName(
            self, "Save CSV", f"sales_datewise_{date.today()}.csv", "CSV Files (*.csv)")
        if not path: return
        try:
            fields = ["date", "invoice_number", "customer_name",
                      "gross_sale", "discount", "sale_amt",
                      "cgst", "sgst", "igst",
                      "advance", "cash", "card", "upi", "cheque", "dues"]
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
                w.writeheader()
                for row in self._rows:
                    w.writerow({k: row.get(k, "") for k in fields})
            QMessageBox.information(self, "Export", f"Exported to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    # ─────────────────────────────────────────────────────────
    #  Export Excel
    # ─────────────────────────────────────────────────────────
    def _export_excel(self):
        if not self._rows:
            QMessageBox.information(self, "Export", "No data to export."); return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(self, "Missing Library",
                "openpyxl is not installed.\n\nRun: pip install openpyxl")
            return

        from PyQt5.QtWidgets import QFileDialog
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel", f"sales_datewise_{date.today()}.xlsx", "Excel Files (*.xlsx)")
        if not path: return

        try:
            wb  = openpyxl.Workbook()
            ws  = wb.active; ws.title = "Date-wise Sales"
            ws.sheet_view.showGridLines = False
            ws.freeze_panes = "A2"

            thin     = Side(style="thin", color="BFBFBF")
            bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)
            ctr      = Alignment(horizontal="center", vertical="center")
            lft      = Alignment(horizontal="left",   vertical="center")
            hdr_fill = PatternFill("solid", fgColor="1F4E79")
            hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
            alt_fill = PatternFill("solid", fgColor="EBF5FB")
            tot_fill = PatternFill("solid", fgColor="D6E4F0")
            tot_font = Font(name="Calibri", bold=True, size=11)
            mfmt     = "#,##0.00"
            _money_c = {4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15}

            xlsx_hdrs = [
                "Date", "Invoice No.", "Customer",
                "Gross Sale ₹", "Discount ₹", "Sale Amt ₹",
                "CGST ₹", "SGST ₹", "IGST ₹",
                "Advance ₹", "Cash ₹", "Card ₹", "UPI ₹", "Cheque ₹", "Dues ₹",
            ]
            for c, h in enumerate(xlsx_hdrs, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = hdr_font; cell.fill = hdr_fill
                cell.alignment = ctr; cell.border = bdr
            ws.row_dimensions[1].height = 24

            for ri, row in enumerate(self._rows, 2):
                alt  = (ri % 2 == 0)
                data = [
                    row["date"],        row["invoice_number"], row["customer_name"],
                    row["gross_sale"],  row["discount"],       row["sale_amt"],
                    row["cgst"],        row["sgst"],           row["igst"],
                    row["advance"],     row["cash"],           row["card"],
                    row["upi"],         row["cheque"],         row["dues"],
                ]
                for c, v in enumerate(data, 1):
                    cell = ws.cell(row=ri, column=c, value=v)
                    cell.alignment = ctr; cell.border = bdr
                    if alt: cell.fill = alt_fill
                    if c in _money_c: cell.number_format = mfmt
                ws.row_dimensions[ri].height = 20

            tr = len(self._rows) + 2
            ws.row_dimensions[tr].height = 22
            for c in range(1, len(xlsx_hdrs) + 1):
                cell = ws.cell(row=tr, column=c)
                cell.fill = tot_fill; cell.border = bdr
                cell.alignment = ctr; cell.font = tot_font
            ws.cell(row=tr, column=1, value="TOTALS")

            totals_map = {
                4:  sum(r["gross_sale"] for r in self._rows),
                5:  sum(r["discount"]   for r in self._rows),
                6:  sum(r["sale_amt"]   for r in self._rows),
                7:  sum(r["cgst"]       for r in self._rows),
                8:  sum(r["sgst"]       for r in self._rows),
                9:  sum(r["igst"]       for r in self._rows),
                10: sum(r["advance"]    for r in self._rows),
                11: sum(r["cash"]       for r in self._rows),
                12: sum(r["card"]       for r in self._rows),
                13: sum(r["upi"]        for r in self._rows),
                14: sum(r["cheque"]     for r in self._rows),
                15: sum(r["dues"]       for r in self._rows),
            }
            for c, v in totals_map.items():
                cell = ws.cell(row=tr, column=c, value=v)
                cell.fill = tot_fill; cell.border = bdr
                cell.alignment = ctr; cell.font = tot_font
                if c in _money_c: cell.number_format = mfmt

            for col in ws.columns:
                w = 8
                for cell in col:
                    try: w = max(w, len(str(cell.value or "")))
                    except: pass
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 30)

            shop = AppConfig.shop()
            ws2 = wb.create_sheet("Summary")
            ws2.sheet_view.showGridLines = False
            ws2.merge_cells("A1:B1")
            t = ws2.cell(1, 1, f"{shop.get('shop_name','Jewelry Shop')}  —  Date-wise Sales")
            t.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
            t.alignment = lft
            ws2.merge_cells("A2:B2")
            p = ws2.cell(2, 1,
                f"Period: {self.dt_from.date().toString('yyyy-MM-dd')}  to  "
                f"{self.dt_to.date().toString('yyyy-MM-dd')}    |    Generated: {date.today()}")
            p.font = Font(name="Calibri", italic=True, color="7F7F7F", size=10)
            p.alignment = lft
            ws2.append([])
            sub_fill = PatternFill("solid", fgColor="D6E4F0")
            _t_sale = sum(r["sale_amt"] for r in self._rows)
            _t_dues = sum(r["dues"]     for r in self._rows)
            for row_label, val, fmt in [
                ("Invoices",                len(self._rows),                              None),
                ("Gross Sale (₹)",        sum(r["gross_sale"] for r in self._rows),  mfmt),
                ("Discount (₹)",          sum(r["discount"]   for r in self._rows),  mfmt),
                ("Sale Amt (₹)",          _t_sale,                                   mfmt),
                ("CGST (₹)",              sum(r["cgst"]       for r in self._rows),  mfmt),
                ("SGST (₹)",              sum(r["sgst"]       for r in self._rows),  mfmt),
                ("IGST (₹)",              sum(r["igst"]       for r in self._rows),  mfmt),
                ("Revenue Collected (₹)", round(_t_sale - _t_dues, 2),               mfmt),
                ("Total Dues (₹)",        _t_dues,                                   mfmt),
                ("  Cash (₹)",            sum(r["cash"]       for r in self._rows),  mfmt),
                ("  Card (₹)",            sum(r["card"]       for r in self._rows),  mfmt),
                ("  UPI (₹)",             sum(r["upi"]        for r in self._rows),  mfmt),
                ("  Cheque (₹)",          sum(r["cheque"]     for r in self._rows),  mfmt),
            ]:
                r_idx = ws2.max_row + 1
                lc = ws2.cell(r_idx, 1, row_label)
                lc.font = Font(name="Calibri", bold=True, size=10)
                lc.fill = sub_fill; lc.alignment = lft; lc.border = bdr
                vc = ws2.cell(r_idx, 2, val)
                vc.alignment = ctr; vc.border = bdr
                if fmt: vc.number_format = fmt
            ws2.column_dimensions["A"].width = 22
            ws2.column_dimensions["B"].width = 18

            wb.save(path)
            QMessageBox.information(self, "Export Successful", f"Excel saved to:\n{path}")
            try: os.startfile(path)
            except: pass

        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))
